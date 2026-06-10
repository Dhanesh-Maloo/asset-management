from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Request, Response, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import smtplib
import asyncio
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, field_validator, ConfigDict
from typing import List, Optional
import uuid
import re
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
from enum import Enum
import httpx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Validate required environment variables on startup
def _validate_env():
    required = ["MONGO_URL", "DB_NAME", "JWT_SECRET"]
    missing = [v for v in required if not os.environ.get(v)]
    if missing:
        raise RuntimeError(f"Missing required environment variables: {', '.join(missing)}")
    weak_secrets = ("your-secret-key-change-in-production", "change-this-to-a-long-random-string", "")
    if os.environ.get("JWT_SECRET", "") in weak_secrets:
        raise RuntimeError("JWT_SECRET is weak or default. Please set a strong random secret in .env")

_validate_env()

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# Email (SMTP) configuration
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
FROM_EMAIL = os.environ.get('FROM_EMAIL', SMTP_USER)
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')

# Google OAuth (direct) configuration
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
GOOGLE_REDIRECT_URI = os.environ.get('GOOGLE_REDIRECT_URI', '')

# Exchange rate cache (refreshed every hour)
_exchange_rates_cache: dict = {"rates": None, "fetched_at": None}

# Enums
class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    TENANT_ADMIN = "tenant_admin"
    ASSET_MANAGER = "asset_manager"
    HELPDESK_AGENT = "helpdesk_agent"
    EMPLOYEE = "employee"

class AssetStatus(str, Enum):
    AVAILABLE = "available"
    ASSIGNED = "assigned"
    IN_USE = "in_use"
    UNDER_MAINTENANCE = "under_maintenance"
    DISPOSED = "disposed"
    CHECKED_OUT = "checked_out"

class MaintenanceStatus(str, Enum):
    SCHEDULED = "scheduled"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    OVERDUE = "overdue"

class DepreciationMethod(str, Enum):
    STRAIGHT_LINE = "straight_line"
    DECLINING_BALANCE = "declining_balance"
    NO_DEPRECIATION = "no_depreciation"

class OrderStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"
    FULFILLED = "fulfilled"

class TicketStatus(str, Enum):
    OPEN = "open"
    IN_PROGRESS = "in_progress"
    RESOLVED = "resolved"
    CLOSED = "closed"

class TicketPriority(str, Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    CRITICAL = "critical"

class GroupType(str, Enum):
    ADMIN_GROUP = "admin_group"
    USER_GROUP = "user_group"

class ApprovalStage(str, Enum):
    MAKER = "maker"
    CHECKER = "checker"
    APPROVER = "approver"

class ApprovalStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"

# Models

class SubscriptionTier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    slug: str
    description: str = ""
    sort_order: int = 0
    is_default: bool = False
    limits: dict = Field(default_factory=lambda: {
        "max_users": 3,
        "max_assets": 10,
        "max_orders_per_month": 5,
        "max_tickets_per_month": 10,
    })
    allowed_features: List[str] = ["products", "orders", "assets", "tickets"]
    highlights: List[str] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SubscriptionTierCreate(BaseModel):
    name: str
    slug: str
    description: str = ""
    sort_order: int = 0
    is_default: bool = False
    limits: dict = Field(default_factory=lambda: {
        "max_users": 3,
        "max_assets": 10,
        "max_orders_per_month": 5,
        "max_tickets_per_month": 10,
    })
    allowed_features: List[str] = ["products", "orders", "assets", "tickets"]
    highlights: List[str] = []

class SubscriptionTierUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    sort_order: Optional[int] = None
    is_default: Optional[bool] = None
    limits: Optional[dict] = None
    allowed_features: Optional[List[str]] = None
    highlights: Optional[List[str]] = None

class Tenant(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    domain: str
    subdomain: str = ""
    custom_domain: str = ""
    logo_url: str = ""
    primary_color: str = "#4F46E5"
    secondary_color: str = "#F1F5F9"
    company_name: str = ""
    enabled_features: List[str] = ["products", "orders", "assets", "tickets", "users", "groups", "workflows"]
    settings: dict = {}
    status: str = "active"
    subscription_tier_id: Optional[str] = None
    subscription_started_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TenantCreate(BaseModel):
    name: str
    domain: str
    subdomain: str = ""
    custom_domain: str = ""
    company_name: str = ""

class TenantUpdate(BaseModel):
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    company_name: Optional[str] = None
    enabled_features: Optional[List[str]] = None
    settings: Optional[dict] = None
    slack_webhook_url: Optional[str] = None

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: UserRole
    tenant_id: Optional[str] = None
    group_id: Optional[str] = None
    department_id: Optional[str] = None
    permissions: List[str] = []
    status: str = "active"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: UserRole
    tenant_id: Optional[str] = None
    group_id: Optional[str] = None

class Group(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    group_type: GroupType
    tenant_id: str
    permissions: List[str] = []
    description: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class GroupCreate(BaseModel):
    name: str
    group_type: GroupType
    tenant_id: str
    permissions: List[str] = []
    description: str = ""

class ApprovalWorkflow(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    entity_type: str
    requires_checker: bool = True
    requires_approver: bool = True
    min_value_threshold: float = 0.0
    checker_group_id: Optional[str] = None
    approver_group_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ApprovalWorkflowCreate(BaseModel):
    tenant_id: str
    entity_type: str
    requires_checker: bool = True
    requires_approver: bool = True
    min_value_threshold: float = 0.0
    checker_group_id: Optional[str] = None
    approver_group_id: Optional[str] = None

class ApprovalHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_id: str
    entity_type: str
    stage: ApprovalStage
    status: ApprovalStatus
    approved_by: str
    comments: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: str
    description: str
    specs: dict = {}
    image_url: str = ""
    stock: int = 0
    price: float = 0.0
    tenant_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ProductCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    category: str = Field(..., min_length=1, max_length=100)
    description: str = Field("", max_length=2000)
    specs: dict = {}
    image_url: str = Field("", max_length=500)
    stock: int = Field(0, ge=0)
    price: float = Field(..., gt=0, description="Price must be greater than 0")
    tenant_id: Optional[str] = None

class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    user_id: str
    product_id: str
    quantity: int
    delivery_notes: str = ""
    status: OrderStatus = OrderStatus.PENDING
    approval_stage: Optional[ApprovalStage] = None
    checker_approved_by: Optional[str] = None
    checker_approved_date: Optional[str] = None
    approver_approved_by: Optional[str] = None
    approver_approved_date: Optional[str] = None
    approved_by: Optional[str] = None
    approval_date: Optional[str] = None
    rejection_reason: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class OrderCreate(BaseModel):
    product_id: str
    quantity: int = 1
    delivery_notes: str = ""

class OrderUpdate(BaseModel):
    status: OrderStatus

class Asset(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_tag: str
    product_id: str
    serial_number: str
    tenant_id: str
    status: AssetStatus = AssetStatus.AVAILABLE
    assigned_to: Optional[str] = None
    assigned_date: Optional[str] = None
    location: str = ""
    
    # Warranty tracking
    purchase_date: Optional[str] = None
    warranty_start_date: Optional[str] = None
    warranty_end_date: Optional[str] = None
    warranty_provider: str = ""
    
    # Depreciation tracking
    purchase_price: float = 0.0
    depreciation_method: DepreciationMethod = DepreciationMethod.STRAIGHT_LINE
    depreciation_rate: float = 20.0
    salvage_value: float = 0.0
    current_value: float = 0.0
    
    # Checkout tracking
    checked_out_to: Optional[str] = None
    checkout_date: Optional[str] = None
    expected_return_date: Optional[str] = None
    actual_return_date: Optional[str] = None

    department_id: Optional[str] = None
    expiry_date: Optional[str] = None
    is_demo: bool = False

    # Lease tracking
    is_leased: bool = False
    lessor_name: str = ""
    lease_start_date: Optional[str] = None
    lease_end_date: Optional[str] = None
    monthly_lease_payment: float = 0.0

    # Disposal tracking
    disposal_date: Optional[str] = None
    disposal_method: str = ""
    disposal_notes: str = ""
    sale_proceeds: float = 0.0

    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class AssetCreate(BaseModel):
    asset_tag: str = Field(..., min_length=1, max_length=100)
    product_id: str
    serial_number: str = Field(..., min_length=1, max_length=100)
    tenant_id: str
    location: str = Field("", max_length=200)
    purchase_date: Optional[str] = None
    warranty_start_date: Optional[str] = None
    warranty_end_date: Optional[str] = None
    warranty_provider: str = ""
    purchase_price: float = 0.0
    depreciation_method: DepreciationMethod = DepreciationMethod.STRAIGHT_LINE
    depreciation_rate: float = 20.0
    salvage_value: float = 0.0
    expiry_date: Optional[str] = None
    is_demo: bool = False
    is_leased: bool = False
    lessor_name: str = ""
    lease_start_date: Optional[str] = None
    lease_end_date: Optional[str] = None
    monthly_lease_payment: float = 0.0

class AssetAssign(BaseModel):
    assigned_to: str

class AssetUpdate(BaseModel):
    status: Optional[AssetStatus] = None
    location: Optional[str] = None
    serial_number: Optional[str] = None
    asset_tag: Optional[str] = None
    purchase_date: Optional[str] = None
    warranty_start_date: Optional[str] = None
    warranty_end_date: Optional[str] = None
    warranty_provider: Optional[str] = None
    purchase_price: Optional[float] = None
    depreciation_method: Optional[DepreciationMethod] = None
    depreciation_rate: Optional[float] = None
    salvage_value: Optional[float] = None
    expiry_date: Optional[str] = None
    department_id: Optional[str] = None
    is_leased: Optional[bool] = None
    lessor_name: Optional[str] = None
    lease_start_date: Optional[str] = None
    lease_end_date: Optional[str] = None
    monthly_lease_payment: Optional[float] = None

class Ticket(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ticket_number: str
    tenant_id: str
    created_by: str
    assigned_to: Optional[str] = None
    title: str
    description: str
    priority: TicketPriority = TicketPriority.MEDIUM
    status: TicketStatus = TicketStatus.OPEN
    category: str = "general"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TicketCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=300)
    description: str = Field(..., min_length=1, max_length=5000)
    priority: TicketPriority = TicketPriority.MEDIUM
    category: str = Field("general", max_length=100)

class TicketUpdate(BaseModel):
    status: Optional[TicketStatus] = None
    assigned_to: Optional[str] = None
    priority: Optional[TicketPriority] = None

class AssetHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_id: str
    action: str
    performed_by: str
    date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    notes: str = ""

class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: str = ""
    tenant_id: str
    budget: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class DepartmentCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    description: str = Field("", max_length=500)
    tenant_id: str
    budget: float = 0.0

class DepartmentUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=100)
    description: Optional[str] = Field(None, max_length=500)
    budget: Optional[float] = None

class MaintenanceSchedule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_id: str
    tenant_id: str
    title: str
    description: str = ""
    scheduled_date: str
    completed_date: Optional[str] = None
    status: MaintenanceStatus = MaintenanceStatus.SCHEDULED
    assigned_to: Optional[str] = None
    maintenance_type: str = "preventive"
    cost: float = 0.0
    notes: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MaintenanceScheduleCreate(BaseModel):
    asset_id: str
    title: str
    description: str = ""
    scheduled_date: str
    assigned_to: Optional[str] = None
    maintenance_type: str = "preventive"

class MaintenanceScheduleUpdate(BaseModel):
    status: Optional[MaintenanceStatus] = None
    completed_date: Optional[str] = None
    cost: Optional[float] = None
    notes: Optional[str] = None

class AssetCheckout(BaseModel):
    checked_out_to: str
    expected_return_date: str

class AssetReturn(BaseModel):
    condition_notes: str = ""
    return_condition: str = "good"  # good, fair, damaged
    daily_fine_rate: float = 0.0   # fine per day for late returns

# ── New Models ──────────────────────────────────────────────────────────────

class Location(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    building: str = ""
    floor: str = ""
    room: str = ""
    tenant_id: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LocationCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    building: str = Field("", max_length=100)
    floor: str = Field("", max_length=50)
    room: str = Field("", max_length=100)
    tenant_id: str

class LocationUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    building: Optional[str] = Field(None, max_length=100)
    floor: Optional[str] = Field(None, max_length=50)
    room: Optional[str] = Field(None, max_length=100)

class Vendor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_name: str = ""
    email: str = ""
    phone: str = ""
    website: str = ""
    category: str = ""
    tenant_id: str
    notes: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class VendorCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    contact_name: str = Field("", max_length=100)
    email: str = Field("", max_length=200)
    phone: str = Field("", max_length=50)
    website: str = Field("", max_length=300)
    category: str = Field("", max_length=100)
    tenant_id: str
    notes: str = Field("", max_length=1000)

class VendorUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    contact_name: Optional[str] = Field(None, max_length=100)
    email: Optional[str] = Field(None, max_length=200)
    phone: Optional[str] = Field(None, max_length=50)
    website: Optional[str] = Field(None, max_length=300)
    category: Optional[str] = Field(None, max_length=100)
    notes: Optional[str] = Field(None, max_length=1000)

class SoftwareLicense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    vendor: str = ""
    license_key: str = ""
    seats_total: int = 1
    seats_used: int = 0
    purchase_date: Optional[str] = None
    expiry_date: Optional[str] = None
    cost: float = 0.0
    license_type: str = "perpetual"
    tenant_id: str
    notes: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LicenseCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    vendor: str = Field("", max_length=200)
    license_key: str = Field("", max_length=500)
    seats_total: int = Field(1, ge=1)
    seats_used: int = Field(0, ge=0)
    purchase_date: Optional[str] = None
    expiry_date: Optional[str] = None
    cost: float = 0.0
    license_type: str = Field("perpetual", max_length=50)
    tenant_id: str
    notes: str = Field("", max_length=1000)

class LicenseUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    vendor: Optional[str] = Field(None, max_length=200)
    seats_total: Optional[int] = None
    seats_used: Optional[int] = None
    expiry_date: Optional[str] = None
    cost: Optional[float] = None
    license_type: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=1000)

class AssetReservation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_id: str
    reserved_by: str
    tenant_id: str
    start_date: str
    end_date: str
    purpose: str = ""
    status: str = "pending"  # pending, approved, rejected, cancelled, completed
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ReservationCreate(BaseModel):
    asset_id: str
    start_date: str
    end_date: str
    purpose: str = Field("", max_length=500)

class ApiKey(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    tenant_id: Optional[str] = None
    name: str
    key: str = Field(default_factory=lambda: "ak_" + str(uuid.uuid4()).replace("-", ""))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    last_used_at: Optional[str] = None
    is_active: bool = True

class ApiKeyCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)

class InviteToken(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    role: str = "employee"
    tenant_id: Optional[str] = None
    invited_by: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    expires_at: str = Field(default_factory=lambda: (datetime.now(timezone.utc) + timedelta(days=7)).isoformat())
    accepted: bool = False

class InviteCreate(BaseModel):
    email: EmailStr
    role: str = "employee"
    tenant_id: Optional[str] = None

class InviteAccept(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    password: str = Field(..., min_length=12)

class TicketComment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ticket_id: str
    author_id: str
    author_name: str = ""
    content: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TicketCommentCreate(BaseModel):
    content: str = Field(..., min_length=1, max_length=2000)

class AssetTransfer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_id: str
    asset_tag: str = ""
    from_user_id: str
    to_user_id: str
    requested_by: str
    tenant_id: str
    reason: str = ""
    status: str = "pending"  # pending, approved, rejected, completed
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TransferCreate(BaseModel):
    asset_id: str
    to_user_id: str
    reason: str = Field("", max_length=500)

class CustomField(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    field_name: str
    field_label: str
    field_type: str = "text"  # text, number, date, select
    options: List[str] = []
    required: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CustomFieldCreate(BaseModel):
    tenant_id: str
    field_name: str = Field(..., min_length=1, max_length=50)
    field_label: str = Field(..., min_length=1, max_length=100)
    field_type: str = Field("text", pattern="^(text|number|date|select)$")
    options: List[str] = []
    required: bool = False

class BulkAssetUpdate(BaseModel):
    ids: List[str]
    status: Optional[AssetStatus] = None
    location: Optional[str] = None
    department_id: Optional[str] = None

class AssetDispose(BaseModel):
    disposal_method: str = Field(..., description="sold, scrapped, donated, or other")
    disposal_date: str
    disposal_notes: str = Field("", max_length=1000)
    sale_proceeds: float = 0.0

class AssetDocument(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_id: str
    tenant_id: str
    document_type: str = "other"  # invoice, warranty, amc, insurance, other
    filename: str
    file_data: str  # base64 data URL
    file_content_type: str = "application/pdf"
    file_size: int = 0
    uploaded_by: str = ""
    notes: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class AMCContract(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    asset_id: str
    tenant_id: str
    vendor_name: str
    contract_number: str = ""
    start_date: str
    end_date: str
    annual_cost: float = 0.0
    renewal_reminder_days: int = 30
    notes: str = ""
    status: str = "active"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class AMCContractCreate(BaseModel):
    vendor_name: str = Field(..., min_length=1, max_length=200)
    contract_number: str = Field("", max_length=100)
    start_date: str
    end_date: str
    annual_cost: float = 0.0
    renewal_reminder_days: int = 30
    notes: str = Field("", max_length=1000)

class AMCContractUpdate(BaseModel):
    vendor_name: Optional[str] = None
    contract_number: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    annual_cost: Optional[float] = None
    renewal_reminder_days: Optional[int] = None
    notes: Optional[str] = None
    status: Optional[str] = None

class ProfileUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=1, max_length=200)

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    amount: float
    currency: str = "USD"
    status: str = "pending"  # pending, paid, overdue
    period_start: str
    period_end: str
    description: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    paid_at: Optional[str] = None

class InvoiceCreate(BaseModel):
    amount: float = Field(..., gt=0)
    currency: str = "USD"
    status: str = "pending"
    period_start: str
    period_end: str
    description: str = Field(..., min_length=1, max_length=300)

class InvoiceUpdate(BaseModel):
    status: Optional[str] = None
    paid_at: Optional[str] = None
    amount: Optional[float] = None

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.JWTError:
        # Try as API key
        token = credentials.credentials
        api_key = await db.api_keys.find_one({"key": token, "is_active": True}, {"_id": 0})
        if api_key:
            user = await db.users.find_one({"id": api_key["user_id"]}, {"_id": 0})
            if user:
                await db.api_keys.update_one(
                    {"key": token},
                    {"$set": {"last_used_at": datetime.now(timezone.utc).isoformat()}}
                )
                return User(**user)
        raise HTTPException(status_code=401, detail="Could not validate credentials")

async def get_current_user_hybrid(request: Request):
    """Support both JWT and OAuth session authentication.
    Priority: Bearer JWT > API key > OAuth session cookie.
    This ensures email/password login always wins over a leftover Google OAuth cookie.
    """
    # ── 1. Try Authorization Bearer header first ─────────────────────────────
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.replace("Bearer ", "")
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                user = await db.users.find_one({"id": user_id}, {"_id": 0})
                if user:
                    return User(**user)
        except jwt.ExpiredSignatureError:
            raise HTTPException(status_code=401, detail="Token has expired")
        except jwt.JWTError:
            # Try as API key
            api_key = await db.api_keys.find_one({"key": token, "is_active": True}, {"_id": 0})
            if api_key:
                user = await db.users.find_one({"id": api_key["user_id"]}, {"_id": 0})
                if user:
                    await db.api_keys.update_one(
                        {"key": token},
                        {"$set": {"last_used_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    return User(**user)

    # ── 2. Fall back to OAuth session cookie ────────────────────────────────
    session_token = request.cookies.get("session_token")
    if session_token:
        session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
        if session:
            expires_at = session["expires_at"]
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at >= datetime.now(timezone.utc):
                user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0})
                if user:
                    return User(**user)
            else:
                await db.user_sessions.delete_one({"session_token": session_token})

    raise HTTPException(status_code=401, detail="Not authenticated")

# ── Helper: send password reset email via SMTP ──────────────────────────────
async def send_reset_email(to_email: str, reset_token: str, user_name: str = "") -> bool:
    """Send a password-reset link via SMTP. Returns True on success."""
    if not SMTP_USER or not SMTP_PASSWORD:
        logging.warning("SMTP not configured — skipping email send (set SMTP_USER/SMTP_PASSWORD in .env)")
        return False

    reset_link = f"{FRONTEND_URL}/reset-password?token={reset_token}"
    greeting = f"Hello {user_name}," if user_name else "Hello,"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = "Password Reset Request"
    msg["From"] = FROM_EMAIL
    msg["To"] = to_email

    html_body = f"""
    <html>
      <body style="font-family:Arial,sans-serif;max-width:600px;margin:auto;padding:20px;">
        <h2 style="color:#4F46E5;">Password Reset Request</h2>
        <p>{greeting}</p>
        <p>We received a request to reset your password for your Asset Management account.</p>
        <p style="margin:24px 0;">
          <a href="{reset_link}"
             style="background:#4F46E5;color:#fff;padding:12px 28px;text-decoration:none;
                    border-radius:6px;display:inline-block;font-weight:bold;">
            Reset Password
          </a>
        </p>
        <p style="color:#6B7280;font-size:14px;">
          Or paste this link into your browser:<br/>
          <a href="{reset_link}" style="color:#4F46E5;">{reset_link}</a>
        </p>
        <p style="color:#6B7280;font-size:13px;">
          This link expires in <strong>1 hour</strong>. If you did not request a password
          reset, you can safely ignore this email.
        </p>
      </body>
    </html>
    """
    msg.attach(MIMEText(html_body, "html"))

    def _send():
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(FROM_EMAIL, to_email, msg.as_string())

    try:
        await asyncio.to_thread(_send)
        logging.info(f"Password reset email sent to {to_email}")
        return True
    except Exception as e:
        logging.error(f"Failed to send reset email to {to_email}: {e}")
        return False


# ── Helper: generic alert email ─────────────────────────────────────────────
async def send_alert_email(to_email: str, subject: str, items: list, intro: str = "") -> bool:
    """Send a generic alert email with a list of items."""
    if not SMTP_USER or not SMTP_PASSWORD:
        logging.warning("SMTP not configured — skipping alert email")
        return False

    rows = "".join(
        f"<tr><td style='padding:8px;border-bottom:1px solid #eee;'>{item}</td></tr>"
        for item in items
    )
    html_body = f"""
    <html>
      <body style="font-family:Arial,sans-serif;max-width:600px;margin:auto;padding:20px;">
        <h2 style="color:#4F46E5;">IT Asset Management — Alert</h2>
        <h3>{subject}</h3>
        {f'<p>{intro}</p>' if intro else ''}
        <table style="width:100%;border-collapse:collapse;margin-top:12px;">
          {rows}
        </table>
        <p style="color:#6B7280;font-size:13px;margin-top:24px;">
          Log in to <a href="{FRONTEND_URL}" style="color:#4F46E5;">IT Asset Management</a> to take action.
        </p>
      </body>
    </html>
    """
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = FROM_EMAIL
    msg["To"] = to_email
    msg.attach(MIMEText(html_body, "html"))

    def _send():
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(FROM_EMAIL, to_email, msg.as_string())

    try:
        await asyncio.to_thread(_send)
        logging.info(f"Alert email '{subject}' sent to {to_email}")
        return True
    except Exception as e:
        logging.error(f"Failed to send alert email: {e}")
        return False

# ── Helper: Slack/Teams webhook notification ────────────────────────────────
async def send_slack_notification(webhook_url: str, title: str, items: list) -> bool:
    """Send a notification to a Slack/Teams incoming webhook."""
    if not webhook_url:
        return False
    text = f"*{title}*\n" + "\n".join(f"• {item}" for item in items[:20])
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(webhook_url, json={"text": text})
            return resp.status_code == 200
    except Exception as e:
        logging.error(f"Slack notification failed: {e}")
        return False

# ── Scheduled background alerts ─────────────────────────────────────────────
async def run_scheduled_alerts():
    """Background task: runs every 24 hours, sends warranty & maintenance alerts."""
    await asyncio.sleep(30)  # Small delay on startup to let DB connect
    while True:
        try:
            logging.info("Running scheduled alert checks...")
            now = datetime.now(timezone.utc)
            soon_30 = (now + timedelta(days=30)).isoformat()
            soon_7  = (now + timedelta(days=7)).isoformat()
            today   = now.isoformat()

            # Gather all tenant admins grouped by tenant
            admins = await db.users.find(
                {"role": {"$in": ["tenant_admin", "super_admin"]}, "status": "active"},
                {"_id": 0, "id": 1, "email": 1, "name": 1, "tenant_id": 1}
            ).to_list(500)

            tenant_admins: dict = {}
            for a in admins:
                tid = a.get("tenant_id") or "__super__"
                tenant_admins.setdefault(tid, []).append(a)

            # Fetch all active tenants for Slack webhook
            tenants_list = await db.tenants.find({}, {"_id": 0, "id": 1, "settings": 1}).to_list(200)
            tenant_slack: dict = {t["id"]: t.get("settings", {}).get("slack_webhook_url", "") for t in tenants_list}

            # ── Warranty expiry alerts ──────────────────────────────────────
            expiring = await db.assets.find(
                {"warranty_end_date": {"$gte": today, "$lte": soon_30}, "is_deleted": {"$ne": True}},
                {"_id": 0, "asset_tag": 1, "name": 1, "warranty_end_date": 1, "tenant_id": 1}
            ).to_list(500)

            warranty_by_tenant: dict = {}
            for a in expiring:
                tid = a.get("tenant_id") or "__super__"
                warranty_by_tenant.setdefault(tid, []).append(
                    f"{a.get('asset_tag','?')} — {a.get('name','?')} (expires {a.get('warranty_end_date','?')[:10]})"
                )

            for tid, items in warranty_by_tenant.items():
                recipients = tenant_admins.get(tid, tenant_admins.get("__super__", []))
                for admin in recipients:
                    await send_alert_email(
                        admin["email"],
                        f"⚠️ {len(items)} Asset(s) Warranty Expiring Soon",
                        items,
                        intro="The following assets have warranties expiring within 30 days:"
                    )
                webhook = tenant_slack.get(tid, "")
                if webhook:
                    await send_slack_notification(webhook, f"⚠️ {len(items)} warranties expiring soon", items)

            # ── Overdue maintenance alerts ──────────────────────────────────
            # Mark overdue tasks in DB so UI shows correct badge
            await db.maintenance_schedules.update_many(
                {"scheduled_date": {"$lt": today}, "status": {"$in": ["scheduled", "in_progress"]}},
                {"$set": {"status": "overdue"}}
            )
            overdue = await db.maintenance_schedules.find(
                {"scheduled_date": {"$lt": today}, "status": "overdue"},
                {"_id": 0, "asset_id": 1, "title": 1, "scheduled_date": 1, "tenant_id": 1}
            ).to_list(500)

            maint_by_tenant: dict = {}
            for m in overdue:
                tid = m.get("tenant_id") or "__super__"
                maint_by_tenant.setdefault(tid, []).append(
                    f"{m.get('title','?')} — scheduled {m.get('scheduled_date','?')[:10]}"
                )

            for tid, items in maint_by_tenant.items():
                recipients = tenant_admins.get(tid, tenant_admins.get("__super__", []))
                for admin in recipients:
                    await send_alert_email(
                        admin["email"],
                        f"🔧 {len(items)} Overdue Maintenance Task(s)",
                        items,
                        intro="The following maintenance tasks are overdue:"
                    )
                webhook = tenant_slack.get(tid, "")
                if webhook:
                    await send_slack_notification(webhook, f"🔧 {len(items)} overdue maintenance tasks", items)

            # ── Weekly summary report (every Monday) ───────────────────────
            if now.weekday() == 0:  # Monday
                all_tenants = await db.tenants.find({"status": {"$ne": "inactive"}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
                for tenant in all_tenants:
                    tid = tenant["id"]
                    t_assets = await db.assets.count_documents({"tenant_id": tid, "is_deleted": {"$ne": True}})
                    t_open_tickets = await db.tickets.count_documents({"tenant_id": tid, "status": {"$in": ["open", "in_progress"]}})
                    t_pending_orders = await db.orders.count_documents({"tenant_id": tid, "status": {"$in": ["pending", "approved"]}})
                    summary_lines = [
                        f"Total assets: {t_assets}",
                        f"Open / In-progress tickets: {t_open_tickets}",
                        f"Pending orders: {t_pending_orders}",
                    ]
                    recipients = tenant_admins.get(tid, [])
                    for admin in recipients:
                        await send_alert_email(
                            admin["email"],
                            f"📊 Weekly Asset Management Summary — {tenant.get('name', tid)}",
                            summary_lines,
                            intro=f"Here is your weekly summary for the week ending {now.strftime('%Y-%m-%d')}:"
                        )

            logging.info("Scheduled alert checks complete.")
        except Exception as e:
            logging.error(f"Scheduled alert error: {e}")

        await asyncio.sleep(86400)  # Sleep 24 hours

# ── Helper: live exchange rates with 1-hour cache ───────────────────────────
async def get_exchange_rates() -> dict:
    """Return USD-based exchange rates. Fetches from exchangerate-api.com, falls back to hardcoded values."""
    global _exchange_rates_cache

    fallback = {
        "USD": 1.0,
        "INR": 83.0,
        "AED": 3.67,
        "SAR": 3.75,
        "EUR": 0.92,
        "GBP": 0.79
    }

    now = datetime.now(timezone.utc)
    cached_at = _exchange_rates_cache.get("fetched_at")
    if (_exchange_rates_cache.get("rates") is not None and
            cached_at is not None and
            (now - cached_at).total_seconds() < 3600):
        return _exchange_rates_cache["rates"]

    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                "https://api.exchangerate-api.com/v4/latest/USD",
                timeout=5.0
            )
            resp.raise_for_status()
            raw_rates = resp.json().get("rates", {})
        rates = {k: raw_rates.get(k, fallback[k]) for k in fallback}
        _exchange_rates_cache = {"rates": rates, "fetched_at": now}
        logging.info("Exchange rates refreshed from live API")
        return rates
    except Exception as e:
        logging.warning(f"Live exchange rate fetch failed ({e}); using fallback rates")
        return fallback


# ── Simple in-memory rate limiter (login / forgot-password) ──────────────────
_rate_limit_store: dict = {}  # {ip: [timestamp, ...]}
MAX_AUTH_ATTEMPTS = 5
AUTH_WINDOW_SECONDS = 300  # 5 minutes

def get_client_ip(request: Request) -> str:
    """Get the real client IP, handling Railway/proxy X-Forwarded-For headers."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    if request.client:
        return request.client.host
    return "unknown"

def check_rate_limit(ip: str):
    now = datetime.now(timezone.utc).timestamp()
    attempts = _rate_limit_store.get(ip, [])
    # Drop attempts outside the window
    attempts = [t for t in attempts if now - t < AUTH_WINDOW_SECONDS]
    if len(attempts) >= MAX_AUTH_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many attempts. Please wait 5 minutes and try again.")
    attempts.append(now)
    _rate_limit_store[ip] = attempts

# ── Audit log helper ──────────────────────────────────────────────────────────
async def write_audit_log(user_id: str, action: str, resource: str, resource_id: str, details: str = ""):
    await db.audit_log.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "action": action,
        "resource": resource,
        "resource_id": resource_id,
        "details": details,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

# ── Password complexity check ─────────────────────────────────────────────────
def validate_password_strength(password: str):
    if len(password) < 12:
        raise HTTPException(status_code=400, detail="Password must be at least 12 characters long")
    if not any(c.isupper() for c in password):
        raise HTTPException(status_code=400, detail="Password must contain at least one uppercase letter")
    if not any(c.islower() for c in password):
        raise HTTPException(status_code=400, detail="Password must contain at least one lowercase letter")
    if not any(c.isdigit() for c in password):
        raise HTTPException(status_code=400, detail="Password must contain at least one number")

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Health check endpoint
@app.get("/health")
async def health_check():
    try:
        await db.command("ping")
        db_status = "connected"
    except Exception:
        db_status = "disconnected"
    return {"status": "ok", "database": db_status}

# Auth endpoints
@api_router.post("/auth/session")
async def create_session_from_oauth(request: Request, response: Response):
    """Exchange session_id from Emergent OAuth for user data and create session"""
    session_id = request.headers.get('X-Session-ID')
    
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing session ID")
    
    # Call Emergent Auth to exchange session_id for user data
    async with httpx.AsyncClient() as client:
        try:
            oauth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id},
                timeout=10.0
            )
            oauth_response.raise_for_status()
            user_data = oauth_response.json()
        except Exception as e:
            raise HTTPException(status_code=401, detail=f"Failed to validate session: {str(e)}")
    
    # Extract user info
    email = user_data.get("email")
    name = user_data.get("name")
    picture = user_data.get("picture", "")
    oauth_session_token = user_data.get("session_token")
    
    if not email or not oauth_session_token:
        raise HTTPException(status_code=400, detail="Invalid user data from OAuth")
    
    # Check if user exists
    existing_user = await db.users.find_one({"email": email}, {"_id": 0})
    
    if existing_user:
        # Update user data if needed
        user_id = existing_user["id"]
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"name": name, "picture": picture}}
        )
    else:
        # Create new user - Auto-signup with tenant creation for new users
        user_id = str(uuid.uuid4())
        
        # Extract domain from email for tenant
        domain = email.split("@")[1]
        company_name = domain.split(".")[0].title()
        
        # Create tenant for new user
        tenant = Tenant(
            name=company_name,
            domain=domain,
            subdomain="",
            company_name=company_name,
            primary_color="#4F46E5",
            secondary_color="#F1F5F9",
            enabled_features=["products", "orders", "assets", "tickets"],
            settings={"currency": "USD"}
        )
        await db.tenants.insert_one(tenant.model_dump())
        
        # Create user as tenant admin
        user = User(
            id=user_id,
            email=email,
            name=name,
            role=UserRole.TENANT_ADMIN,
            tenant_id=tenant.id,
            permissions=[],
            status="active"
        )
        user_doc = user.model_dump()
        user_doc["picture"] = picture
        await db.users.insert_one(user_doc)
    
    # Store session in database
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    session_doc = {
        "user_id": user_id,
        "session_token": oauth_session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    }
    await db.user_sessions.insert_one(session_doc)
    
    # Set httpOnly cookie
    response.set_cookie(
        key="session_token",
        value=oauth_session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    # Get complete user data
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    
    # Generate a JWT token so AuthContext can use it seamlessly
    jwt_token = create_access_token({"sub": user_id, "email": email, "role": user["role"]})
    
    return {"message": "Session created", "user": user, "token": jwt_token}


# ── Direct Google OAuth endpoint ─────────────────────────────────────────────
class GoogleAuthRequest(BaseModel):
    code: str
    redirect_uri: Optional[str] = None

@api_router.post("/auth/google")
async def google_oauth_callback(request: GoogleAuthRequest):
    """Exchange a Google OAuth authorization code for a JWT session token."""
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise HTTPException(
            status_code=501,
            detail="Google OAuth is not configured. Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET in environment variables."
        )

    # Use redirect_uri from the request (must match exactly what was used in the
    # authorization request). Fall back to GOOGLE_REDIRECT_URI env var only if
    # the frontend doesn't send one.
    redirect_uri = request.redirect_uri or GOOGLE_REDIRECT_URI
    if not redirect_uri:
        raise HTTPException(status_code=400, detail="No redirect_uri provided.")

    async with httpx.AsyncClient() as client:
        # Step 1: Exchange authorization code for Google tokens
        token_resp = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": request.code,
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": redirect_uri,
                "grant_type": "authorization_code",
            },
        )
        if token_resp.status_code != 200:
            google_error = token_resp.json().get("error_description") or token_resp.json().get("error") or "unknown error"
            logging.error(f"Google token exchange failed: {token_resp.status_code} {google_error}")
            raise HTTPException(status_code=401, detail=f"Google sign-in failed: {google_error}")

        access_token = token_resp.json().get("access_token")

        # Step 2: Retrieve user profile from Google
        userinfo_resp = await client.get(
            "https://www.googleapis.com/oauth2/v1/userinfo",
            headers={"Authorization": f"Bearer {access_token}"},
        )
        if userinfo_resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Failed to retrieve Google user info")

        user_data = userinfo_resp.json()

    email = user_data.get("email")
    name = user_data.get("name", email)
    picture = user_data.get("picture", "")

    if not email:
        raise HTTPException(status_code=400, detail="Google account returned no email address")

    # Check if user already exists; create them if not
    existing_user = await db.users.find_one({"email": email}, {"_id": 0})

    if existing_user:
        user_id = existing_user["id"]
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"name": name, "picture": picture}},
        )
    else:
        user_id = str(uuid.uuid4())
        domain = email.split("@")[1]
        company_name = domain.split(".")[0].title()

        tenant = Tenant(
            name=company_name,
            domain=domain,
            subdomain="",
            company_name=company_name,
            primary_color="#4F46E5",
            secondary_color="#F1F5F9",
            enabled_features=["products", "orders", "assets", "tickets"],
            settings={"currency": "USD"},
        )
        await db.tenants.insert_one(tenant.model_dump())

        new_user = User(
            id=user_id,
            email=email,
            name=name,
            role=UserRole.TENANT_ADMIN,
            tenant_id=tenant.id,
            permissions=[],
            status="active",
        )
        user_doc = new_user.model_dump()
        user_doc["picture"] = picture
        await db.users.insert_one(user_doc)

    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    jwt_token = create_access_token({"sub": user_id, "email": email, "role": user["role"]})

    return {"message": "Authenticated via Google", "user": user, "token": jwt_token}


async def get_current_user_from_session(request: Request):
    """Get current user from session_token cookie or Authorization header"""
    session_token = request.cookies.get("session_token")
    
    # Fallback to Authorization header
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.replace("Bearer ", "")
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Find session in database
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    # Check expiry
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        await db.user_sessions.delete_one({"session_token": session_token})
        raise HTTPException(status_code=401, detail="Session expired")
    
    # Get user
    user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return User(**user)

@api_router.get("/auth/me", response_model=User)
async def get_me_oauth(request: Request, current_user: User = Depends(get_current_user_hybrid)):
    """Get current user info - works with both OAuth session and JWT"""
    return current_user

@api_router.post("/auth/logout")
async def logout_oauth(request: Request, response: Response):
    """Logout and clear session"""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    if user_data.tenant_id:
        await check_tier_limit(user_data.tenant_id, "users")
    
    hashed_password = hash_password(user_data.password)
    user_dict = user_data.model_dump()
    user_dict.pop("password")
    user = User(**user_dict)
    
    doc = user.model_dump()
    doc["password_hash"] = hashed_password
    await db.users.insert_one(doc)
    
    return {"message": "User registered successfully", "user": user}

@api_router.post("/auth/signup")
async def public_signup(
    company_name: str,
    admin_name: str,
    admin_email: EmailStr,
    admin_password: str,
    domain: str,
    subdomain: str = "",
    currency: str = "USD",
    seed_demo_data: bool = False,
):
    # Input length validation
    if not company_name or len(company_name.strip()) < 2 or len(company_name) > 100:
        raise HTTPException(status_code=400, detail="Company name must be between 2 and 100 characters")
    if not admin_name or len(admin_name.strip()) < 2 or len(admin_name) > 100:
        raise HTTPException(status_code=400, detail="Name must be between 2 and 100 characters")
    if not domain or len(domain) > 253 or not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9\-\.]{0,251}[a-zA-Z0-9]$', domain):
        raise HTTPException(status_code=400, detail="Invalid domain format")
    if subdomain and (len(subdomain) > 63 or not re.match(r'^[a-z0-9][a-z0-9\-]{0,61}[a-z0-9]?$', subdomain.lower())):
        raise HTTPException(status_code=400, detail="Subdomain must be lowercase letters, numbers, and hyphens only (max 63 chars)")
    if currency not in ("USD", "EUR", "GBP", "INR", "AED", "SGD", "AUD", "CAD"):
        raise HTTPException(status_code=400, detail="Unsupported currency")

    # Password strength
    validate_password_strength(admin_password)

    # Check if email already exists
    existing_user = await db.users.find_one({"email": admin_email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Check if subdomain already exists
    if subdomain:
        existing_tenant = await db.tenants.find_one({"subdomain": subdomain}, {"_id": 0})
        if existing_tenant:
            raise HTTPException(status_code=400, detail="Subdomain already taken")
    
    # Create tenant
    tenant = Tenant(
        name=company_name,
        domain=domain,
        subdomain=subdomain.lower(),
        company_name=company_name,
        primary_color="#4F46E5",
        secondary_color="#F1F5F9",
        enabled_features=["products", "orders", "assets", "tickets"],
        settings={"currency": currency}
    )
    await db.tenants.insert_one(tenant.model_dump())
    
    # Create admin user
    hashed_password = hash_password(admin_password)
    admin_user = User(
        email=admin_email,
        name=admin_name,
        role=UserRole.TENANT_ADMIN,
        tenant_id=tenant.id,
        status="active"
    )
    
    admin_doc = admin_user.model_dump()
    admin_doc["password_hash"] = hashed_password
    await db.users.insert_one(admin_doc)
    
    # Seed demo data only if explicitly requested
    if seed_demo_data:
        try:
            await seed_tenant_demo_data(tenant.id, admin_user.id)
        except Exception as e:
            logger.warning(f"Demo data seeding failed for tenant {tenant.id}: {e}")

    return {
        "message": "Signup successful! You can now login.",
        "tenant_id": tenant.id,
        "user_id": admin_user.id
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin, request: Request, response: Response):
    client_ip = get_client_ip(request)
    check_rate_limit(client_ip)
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Clear any stale Google OAuth session cookie so the new JWT takes full control
    old_session = request.cookies.get("session_token")
    if old_session:
        await db.user_sessions.delete_one({"session_token": old_session})
    response.delete_cookie("session_token", path="/")

    user.pop("password_hash", None)
    token = create_access_token({"sub": user["id"], "email": user["email"], "role": user["role"]})
    await write_audit_log(user["id"], "login", "user", user["id"], f"Login from {client_ip}")
    return {"token": token, "user": User(**user)}

# Password Reset endpoints
class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest, req: Request):
    check_rate_limit(get_client_ip(req))
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user:
        # Don't reveal whether email exists
        return {"message": "If an account with this email exists, a password reset link has been sent."}
    
    reset_token = str(uuid.uuid4())
    expires_at = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
    
    await db.password_resets.insert_one({
        "token": reset_token,
        "user_id": user["id"],
        "email": request.email,
        "expires_at": expires_at,
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    # Send real email (or log a warning if SMTP is not configured)
    email_sent = await send_reset_email(
        to_email=request.email,
        reset_token=reset_token,
        user_name=user.get("name", "")
    )

    response: dict = {"message": "If an account with this email exists, a password reset link has been sent."}
    if not email_sent:
        # SMTP not configured — surface the token so the UI can show it directly
        response["reset_token"] = reset_token
    return response

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    reset_record = await db.password_resets.find_one(
        {"token": request.token, "used": False}, {"_id": 0}
    )
    if not reset_record:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
    
    if datetime.now(timezone.utc).isoformat() > reset_record["expires_at"]:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    validate_password_strength(request.new_password)
    
    hashed = hash_password(request.new_password)
    await db.users.update_one({"id": reset_record["user_id"]}, {"$set": {"password_hash": hashed}})
    await db.password_resets.update_one({"token": request.token}, {"$set": {"used": True}})
    
    return {"message": "Password has been reset successfully"}

# Tenant endpoints
@api_router.post("/tenants", response_model=Tenant)
async def create_tenant(tenant_data: TenantCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can create tenants")
    
    tenant = Tenant(**tenant_data.model_dump())
    await db.tenants.insert_one(tenant.model_dump())
    return tenant

@api_router.get("/tenants", response_model=List[Tenant])
async def get_tenants(current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Access denied")
    
    tenants = await db.tenants.find({}, {"_id": 0}).to_list(1000)
    return [Tenant(**t) for t in tenants]

@api_router.get("/tenants/{tenant_id}", response_model=Tenant)
async def get_tenant(tenant_id: str, current_user: User = Depends(get_current_user)):
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    if current_user.role == UserRole.TENANT_ADMIN and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return Tenant(**tenant)

@api_router.patch("/tenants/{tenant_id}", response_model=Tenant)
async def update_tenant(tenant_id: str, tenant_update: TenantUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    if current_user.role == UserRole.TENANT_ADMIN and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_data = tenant_update.model_dump(exclude_unset=True)
    # Handle slack_webhook_url → store in settings dict
    if "slack_webhook_url" in update_data:
        slack_url = update_data.pop("slack_webhook_url")
        if slack_url is not None:
            update_data["settings.slack_webhook_url"] = slack_url
    await db.tenants.update_one({"id": tenant_id}, {"$set": update_data})
    
    updated_tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    return Tenant(**updated_tenant)

@api_router.delete("/tenants/{tenant_id}")
async def delete_tenant(tenant_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can delete tenants")

    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    # Cascade delete all data belonging to this tenant
    for col in ["users", "assets", "tickets", "orders", "products", "departments",
                "groups", "locations", "vendors", "licenses", "reservations",
                "maintenance_tasks", "api_keys", "audit_logs", "custom_fields",
                "asset_transfers", "ticket_comments"]:
        await db[col].delete_many({"tenant_id": tenant_id})

    await db.tenants.delete_one({"id": tenant_id})
    return {"message": f"Tenant '{tenant['name']}' and all related data deleted"}

@api_router.get("/tenants/subdomain/{subdomain}")
async def get_tenant_by_subdomain(subdomain: str):
    tenant = await db.tenants.find_one({"subdomain": subdomain}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    # Return only branding fields needed for the login page — never expose settings or internal IDs
    return {
        "name": tenant.get("name", ""),
        "company_name": tenant.get("company_name", ""),
        "logo_url": tenant.get("logo_url", ""),
        "primary_color": tenant.get("primary_color", "#4F46E5"),
        "secondary_color": tenant.get("secondary_color", "#F1F5F9"),
        "subdomain": tenant.get("subdomain", ""),
    }

# Product endpoints
@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    product_dict = product_data.model_dump()
    # Assign tenant_id: super_admin can create global products (no tenant_id) or tenant-specific
    if current_user.role != UserRole.SUPER_ADMIN:
        product_dict["tenant_id"] = current_user.tenant_id
    elif not product_dict.get("tenant_id"):
        product_dict["tenant_id"] = None  # Global product

    # Case-insensitive duplicate check within the same tenant scope
    scope_tenant_id = product_dict.get("tenant_id")
    duplicate_query = {
        "name": {"$regex": f"^{re.escape(product_dict['name'])}$", "$options": "i"},
        "is_deleted": {"$ne": True},
    }
    if scope_tenant_id:
        duplicate_query["$or"] = [{"tenant_id": scope_tenant_id}, {"tenant_id": None}, {"tenant_id": {"$exists": False}}]
    existing = await db.products.find_one(duplicate_query, {"_id": 0, "name": 1})
    if existing:
        raise HTTPException(status_code=409, detail=f"A product named '{existing['name']}' already exists. Please use a different name.")

    product = Product(**product_dict)
    await db.products.insert_one(product.model_dump())
    return product

@api_router.get("/products")
async def get_products(
    currency: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    search: Optional[str] = Query(None, description="Search by name or category"),
    category: Optional[str] = Query(None, description="Filter by category"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500)
):
    base: dict = {"is_deleted": {"$ne": True}}
    # Filter: show tenant's own products + global products (tenant_id is None)
    if current_user.role == UserRole.SUPER_ADMIN:
        query = base
    else:
        query = {**base, "$or": [
            {"tenant_id": current_user.tenant_id},
            {"tenant_id": None},
            {"tenant_id": {"$exists": False}}
        ]}

    if search:
        query["$and"] = [{"$or": [
            {"name": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}}
        ]}]
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    skip = (page - 1) * limit
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    # Fetch live exchange rates (cached for 1 hour, falls back to hardcoded values)
    conversion_rates = await get_exchange_rates()

    # Get tenant currency if not specified
    if not currency and current_user.tenant_id:
        tenant = await db.tenants.find_one({"id": current_user.tenant_id}, {"_id": 0})
        if tenant and tenant.get("settings", {}).get("currency"):
            currency = tenant["settings"]["currency"]

    # Apply currency conversion
    if currency and currency != "USD":
        rate = conversion_rates.get(currency, 1.0)
        for product in products:
            product["price"] = round(product["price"] * rate, 2)
            product["display_currency"] = currency
    else:
        for product in products:
            product["display_currency"] = "USD"
    
    return products

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return Product(**product)

@api_router.post("/products/bulk-import")
async def bulk_import_products(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Unsupported file format. Please upload a CSV or Excel (.xlsx) file.")
    
    content = await file.read()
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Maximum allowed size is 10 MB.")

    try:
        import io
        import csv as csv_module
        
        rows = []
        
        if filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv_module.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [h.strip().lower().replace(' ', '_') if h else '' for h in headers]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = val
                if any(row_dict.values()):
                    rows.append(row_dict)
            wb.close()
        
        if not rows:
            raise HTTPException(status_code=400, detail="File is empty or has no data rows")
        
        # Map and validate products
        imported = 0
        errors = []
        
        for idx, row in enumerate(rows, start=2):
            name = str(row.get('name', '') or '').strip()
            category = str(row.get('category', '') or '').strip()
            description = str(row.get('description', '') or '').strip()
            
            if not name:
                errors.append(f"Row {idx}: Missing product name")
                continue
            if not category:
                errors.append(f"Row {idx}: Missing category")
                continue
            
            try:
                price = float(row.get('price', 0) or 0)
            except (ValueError, TypeError):
                errors.append(f"Row {idx}: Invalid price")
                continue
            
            if price <= 0:
                errors.append(f"Row {idx}: Price must be greater than 0")
                continue
            
            try:
                stock = int(float(row.get('stock', 0) or 0))
            except (ValueError, TypeError):
                stock = 0
            
            product = Product(
                name=name,
                category=category,
                description=description or name,
                price=price,
                stock=max(0, stock),
                image_url=str(row.get('image_url', '') or '').strip(),
                tenant_id=current_user.tenant_id if current_user.role != UserRole.SUPER_ADMIN else None
            )
            await db.products.insert_one(product.model_dump())
            imported += 1
        
        result = {"imported": imported, "total_rows": len(rows), "errors": errors[:20]}
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

# Order endpoints
@api_router.post("/orders", response_model=Order)
async def create_order(order_data: OrderCreate, current_user: User = Depends(get_current_user)):
    await check_tier_limit(current_user.tenant_id, "orders")
    order = Order(
        **order_data.model_dump(),
        tenant_id=current_user.tenant_id or "default",
        user_id=current_user.id
    )
    await db.orders.insert_one(order.model_dump())

    # Notify admins/managers about new order awaiting approval
    product = await db.products.find_one({"id": order_data.product_id}, {"_id": 0, "name": 1})
    product_name = product["name"] if product else "Unknown Product"
    admin_query = {"role": {"$in": ["tenant_admin", "asset_manager"]}, "tenant_id": current_user.tenant_id}
    admins = await db.users.find(admin_query, {"_id": 0, "email": 1}).to_list(50)
    for admin in admins:
        if admin.get("email"):
            await send_alert_email(
                admin["email"],
                "📦 New Order Awaiting Your Approval",
                [f"{current_user.name} ordered {order_data.quantity}x {product_name}"],
                "A new order has been submitted and is waiting for approval:"
            )
    return order

@api_router.get("/orders", response_model=List[Order])
async def get_orders(
    current_user: User = Depends(get_current_user),
    status: Optional[str] = Query(None, description="Filter by status"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500)
):
    query: dict = {}
    if current_user.role == UserRole.EMPLOYEE:
        query["user_id"] = current_user.id
    elif current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        query["tenant_id"] = current_user.tenant_id

    if status:
        query["status"] = status

    skip = (page - 1) * limit
    orders = await db.orders.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return [Order(**o) for o in orders]

@api_router.patch("/orders/{order_id}", response_model=Order)
async def update_order(order_id: str, order_update: OrderUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    update_data = order_update.model_dump(exclude_unset=True)
    if order_update.status in [OrderStatus.APPROVED, OrderStatus.REJECTED]:
        update_data["approved_by"] = current_user.id
        update_data["approval_date"] = datetime.now(timezone.utc).isoformat()

    # Atomically deduct stock before marking approved — prevents race condition
    if order_update.status == OrderStatus.APPROVED:
        qty = order.get("quantity", 1)
        stock_result = await db.products.find_one_and_update(
            {"id": order["product_id"], "stock": {"$gte": qty}},
            {"$inc": {"stock": -qty}},
            return_document=True
        )
        if stock_result is None:
            raise HTTPException(status_code=409, detail="Insufficient stock — order cannot be approved")

    await db.orders.update_one({"id": order_id}, {"$set": update_data})

    # Notify order creator of approval/rejection
    if order_update.status in [OrderStatus.APPROVED, OrderStatus.REJECTED]:
        creator = await db.users.find_one({"id": order.get("user_id")}, {"_id": 0, "email": 1, "name": 1})
        if creator and creator.get("email"):
            product = await db.products.find_one({"id": order["product_id"]}, {"_id": 0, "name": 1})
            product_name = product["name"] if product else "your item"
            status_word = "approved ✅" if order_update.status == OrderStatus.APPROVED else "rejected ❌"
            await send_alert_email(
                creator["email"],
                f"Order {status_word.upper()} — {product_name}",
                [f"Your order for {order.get('quantity', 1)}x {product_name} has been {status_word} by {current_user.name}."],
                "Here is an update on your order:"
            )

    updated_order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return Order(**updated_order)

# Asset endpoints
@api_router.post("/assets", response_model=Asset)
async def create_asset(asset_data: AssetCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    # Prevent tenant admins/managers from creating assets for other tenants
    if current_user.role != UserRole.SUPER_ADMIN and asset_data.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Cannot create assets for other tenants")

    await check_tier_limit(asset_data.tenant_id, "assets")
    asset = Asset(**asset_data.model_dump())
    await db.assets.insert_one(asset.model_dump())
    
    history = AssetHistory(
        asset_id=asset.id,
        action="created",
        performed_by=current_user.id,
        notes="Asset created"
    )
    await db.asset_history.insert_one(history.model_dump())
    
    return asset

@api_router.get("/assets", response_model=List[Asset])
async def get_assets(
    current_user: User = Depends(get_current_user),
    status: Optional[str] = Query(None, description="Filter by status"),
    search: Optional[str] = Query(None, description="Search by name or serial number"),
    is_demo: Optional[bool] = Query(None, description="Filter demo assets"),
    department_id: Optional[str] = Query(None, description="Filter by department"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(50, ge=1, le=500, description="Results per page")
):
    query: dict = {"is_deleted": {"$ne": True}}
    if current_user.role == UserRole.EMPLOYEE:
        # Employees can see available assets (for reservations) or their own assigned assets
        if status == "available":
            query["tenant_id"] = current_user.tenant_id
        else:
            query["assigned_to"] = current_user.id
    elif current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        query["tenant_id"] = current_user.tenant_id

    if status:
        query["status"] = status
    if is_demo is not None:
        query["is_demo"] = is_demo
    if department_id:
        query["department_id"] = department_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"serial_number": {"$regex": search, "$options": "i"}},
            {"asset_tag": {"$regex": search, "$options": "i"}}
        ]

    skip = (page - 1) * limit
    assets = await db.assets.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return [Asset(**a) for a in assets]

@api_router.patch("/assets/{asset_id}/assign", response_model=Asset)
async def assign_asset(asset_id: str, assignment: AssetAssign, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    update_data = {
        "assigned_to": assignment.assigned_to,
        "assigned_date": datetime.now(timezone.utc).isoformat(),
        "status": AssetStatus.ASSIGNED
    }
    
    await db.assets.update_one({"id": asset_id}, {"$set": update_data})
    
    history = AssetHistory(
        asset_id=asset_id,
        action="assigned",
        performed_by=current_user.id,
        notes=f"Assigned to {assignment.assigned_to}"
    )
    await db.asset_history.insert_one(history.model_dump())
    
    updated_asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    return Asset(**updated_asset)

@api_router.patch("/assets/{asset_id}", response_model=Asset)
async def update_asset(asset_id: str, asset_update: AssetUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    update_data = asset_update.model_dump(exclude_unset=True)
    await db.assets.update_one({"id": asset_id}, {"$set": update_data})
    
    action = "updated"
    if asset_update.status == AssetStatus.DISPOSED:
        action = "disposed"
    elif asset_update.status == AssetStatus.AVAILABLE:
        action = "retrieved"
    
    history = AssetHistory(
        asset_id=asset_id,
        action=action,
        performed_by=current_user.id,
        notes=f"Status changed to {asset_update.status}"
    )
    await db.asset_history.insert_one(history.model_dump())
    
    updated_asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    return Asset(**updated_asset)

# Ticket endpoints
@api_router.post("/tickets", response_model=Ticket)
async def create_ticket(ticket_data: TicketCreate, current_user: User = Depends(get_current_user)):
    await check_tier_limit(current_user.tenant_id, "tickets")
    ticket_count = await db.tickets.count_documents({"tenant_id": current_user.tenant_id or "default"})
    ticket_number = f"TKT-{ticket_count + 1:05d}"
    
    ticket = Ticket(
        **ticket_data.model_dump(),
        ticket_number=ticket_number,
        tenant_id=current_user.tenant_id or "default",
        created_by=current_user.id
    )
    await db.tickets.insert_one(ticket.model_dump())
    return ticket

@api_router.get("/tickets", response_model=List[Ticket])
async def get_tickets(
    current_user: User = Depends(get_current_user),
    status: Optional[str] = Query(None, description="Filter by status"),
    priority: Optional[str] = Query(None, description="Filter by priority"),
    search: Optional[str] = Query(None, description="Search by title"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500)
):
    query: dict = {}
    if current_user.role == UserRole.EMPLOYEE:
        query["created_by"] = current_user.id
    elif current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        query["tenant_id"] = current_user.tenant_id

    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if search:
        query["title"] = {"$regex": search, "$options": "i"}

    skip = (page - 1) * limit
    tickets = await db.tickets.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return [Ticket(**t) for t in tickets]

@api_router.patch("/tickets/{ticket_id}", response_model=Ticket)
async def update_ticket(ticket_id: str, ticket_update: TicketUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    update_data = ticket_update.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.tickets.update_one({"id": ticket_id}, {"$set": update_data})
    updated_ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    return Ticket(**updated_ticket)

# Dashboard stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        query["tenant_id"] = current_user.tenant_id
    
    total_assets = await db.assets.count_documents(query)
    assigned_assets = await db.assets.count_documents({**query, "status": AssetStatus.ASSIGNED})
    available_assets = await db.assets.count_documents({**query, "status": AssetStatus.AVAILABLE})
    
    ticket_query = query.copy()
    if current_user.role == UserRole.EMPLOYEE:
        ticket_query["created_by"] = current_user.id
    
    open_tickets = await db.tickets.count_documents({**ticket_query, "status": TicketStatus.OPEN})
    pending_orders = await db.orders.count_documents({**query, "status": OrderStatus.PENDING})
    
    stats = {
        "total_assets": total_assets,
        "assigned_assets": assigned_assets,
        "available_assets": available_assets,
        "open_tickets": open_tickets,
        "pending_orders": pending_orders
    }
    
    if current_user.role == UserRole.SUPER_ADMIN:
        stats["total_tenants"] = await db.tenants.count_documents({})
        stats["total_users"] = await db.users.count_documents({})
    
    return stats

# Users endpoint
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        raise HTTPException(status_code=403, detail="Access denied")

    query = {}
    if current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        query["tenant_id"] = current_user.tenant_id
    
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [User(**u) for u in users]

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")

    target_user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    if current_user.role == UserRole.TENANT_ADMIN:
        if target_user["tenant_id"] != current_user.tenant_id:
            raise HTTPException(status_code=403, detail="Access denied")
        if target_user["role"] in ["tenant_admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Cannot delete admin accounts")

    await db.users.delete_one({"id": user_id})
    return {"message": "User deleted successfully"}

# Group endpoints
@api_router.post("/groups", response_model=Group)
async def create_group(group_data: GroupCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if current_user.role == UserRole.TENANT_ADMIN and group_data.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Cannot create groups for other tenants")
    
    group = Group(**group_data.model_dump())
    await db.groups.insert_one(group.model_dump())
    return group

@api_router.get("/groups", response_model=List[Group])
async def get_groups(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if current_user.role == UserRole.TENANT_ADMIN:
        query["tenant_id"] = current_user.tenant_id
    
    groups = await db.groups.find(query, {"_id": 0}).to_list(1000)
    return [Group(**g) for g in groups]

@api_router.get("/groups/{group_id}", response_model=Group)
async def get_group(group_id: str, current_user: User = Depends(get_current_user)):
    group = await db.groups.find_one({"id": group_id}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    if current_user.role == UserRole.TENANT_ADMIN and group["tenant_id"] != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return Group(**group)

@api_router.patch("/groups/{group_id}", response_model=Group)
async def update_group(group_id: str, permissions: List[str], current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    group = await db.groups.find_one({"id": group_id}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    if current_user.role == UserRole.TENANT_ADMIN and group["tenant_id"] != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.groups.update_one({"id": group_id}, {"$set": {"permissions": permissions}})
    updated_group = await db.groups.find_one({"id": group_id}, {"_id": 0})
    return Group(**updated_group)

@api_router.delete("/groups/{group_id}")
async def delete_group(group_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    group = await db.groups.find_one({"id": group_id}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    if current_user.role == UserRole.TENANT_ADMIN and group["tenant_id"] != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Check if any users are assigned to this group
    user_count = await db.users.count_documents({"group_id": group_id})
    if user_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete group: {user_count} user(s) are assigned to it. Reassign them first.")
    
    await db.groups.delete_one({"id": group_id})
    return {"message": "Group deleted"}

# Approval Workflow endpoints
@api_router.post("/approval-workflows", response_model=ApprovalWorkflow)
async def create_approval_workflow(workflow_data: ApprovalWorkflowCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if current_user.role == UserRole.TENANT_ADMIN and workflow_data.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Cannot create workflows for other tenants")
    
    workflow = ApprovalWorkflow(**workflow_data.model_dump())
    await db.approval_workflows.insert_one(workflow.model_dump())
    return workflow

@api_router.get("/approval-workflows", response_model=List[ApprovalWorkflow])
async def get_approval_workflows(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        query["tenant_id"] = current_user.tenant_id
    
    workflows = await db.approval_workflows.find(query, {"_id": 0}).to_list(1000)
    return [ApprovalWorkflow(**w) for w in workflows]

# Enhanced order approval with maker-checker-approver
@api_router.post("/orders/{order_id}/checker-approve")
async def checker_approve_order(order_id: str, comments: str = "", current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Validate: order must be in pending stage before checker can act
    if order.get("status") != OrderStatus.PENDING:
        raise HTTPException(status_code=400, detail="Only pending orders can be checker-approved")
    if order.get("approval_stage") == ApprovalStage.CHECKER:
        raise HTTPException(status_code=400, detail="Order already has checker approval")

    update_data = {
        "approval_stage": ApprovalStage.CHECKER,
        "checker_approved_by": current_user.id,
        "checker_approved_date": datetime.now(timezone.utc).isoformat()
    }
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # Log approval history
    history = ApprovalHistory(
        entity_id=order_id,
        entity_type="order",
        stage=ApprovalStage.CHECKER,
        status=ApprovalStatus.APPROVED,
        approved_by=current_user.id,
        comments=comments
    )
    await db.approval_history.insert_one(history.model_dump())
    
    updated_order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return Order(**updated_order)

@api_router.post("/orders/{order_id}/approver-approve")
async def approver_approve_order(order_id: str, comments: str = "", current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Validate stage: checker must have approved before approver can act
    workflow = await db.approval_workflows.find_one(
        {"tenant_id": order.get("tenant_id"), "entity_type": "order"}, {"_id": 0}
    )
    if workflow and workflow.get("requires_checker") and order.get("approval_stage") != ApprovalStage.CHECKER:
        raise HTTPException(status_code=400, detail="Checker approval required before final approval")
    if order.get("status") == OrderStatus.APPROVED:
        raise HTTPException(status_code=400, detail="Order is already approved")
    if order.get("status") == OrderStatus.REJECTED:
        raise HTTPException(status_code=400, detail="Cannot approve a rejected order")

    qty = order.get("quantity", 1)
    stock_result = await db.products.find_one_and_update(
        {"id": order["product_id"], "stock": {"$gte": qty}},
        {"$inc": {"stock": -qty}},
        return_document=True
    )
    if stock_result is None:
        raise HTTPException(status_code=409, detail="Insufficient stock — order cannot be approved")

    update_data = {
        "status": OrderStatus.APPROVED,
        "approval_stage": ApprovalStage.APPROVER,
        "approver_approved_by": current_user.id,
        "approver_approved_date": datetime.now(timezone.utc).isoformat(),
        "approved_by": current_user.id,
        "approval_date": datetime.now(timezone.utc).isoformat()
    }

    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # Log approval history
    history = ApprovalHistory(
        entity_id=order_id,
        entity_type="order",
        stage=ApprovalStage.APPROVER,
        status=ApprovalStatus.APPROVED,
        approved_by=current_user.id,
        comments=comments
    )
    await db.approval_history.insert_one(history.model_dump())
    
    updated_order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return Order(**updated_order)

@api_router.post("/orders/{order_id}/reject")
async def reject_order(order_id: str, rejection_reason: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    update_data = {
        "status": OrderStatus.REJECTED,
        "rejection_reason": rejection_reason
    }
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # Log approval history
    stage = order.get("approval_stage") or ApprovalStage.MAKER
    history = ApprovalHistory(
        entity_id=order_id,
        entity_type="order",
        stage=stage,
        status=ApprovalStatus.REJECTED,
        approved_by=current_user.id,
        comments=rejection_reason
    )
    await db.approval_history.insert_one(history.model_dump())
    
    updated_order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return Order(**updated_order)

@api_router.get("/approval-history/{entity_id}")
async def get_approval_history(entity_id: str, current_user: User = Depends(get_current_user)):
    history = await db.approval_history.find({"entity_id": entity_id}, {"_id": 0}).to_list(100)
    return [ApprovalHistory(**h) for h in history]

# Maintenance Schedule endpoints
@api_router.post("/maintenance", response_model=MaintenanceSchedule)
async def create_maintenance(maintenance_data: MaintenanceScheduleCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    asset = await db.assets.find_one({"id": maintenance_data.asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    maintenance = MaintenanceSchedule(
        **maintenance_data.model_dump(),
        tenant_id=asset["tenant_id"]
    )
    await db.maintenance_schedules.insert_one(maintenance.model_dump())
    return maintenance

@api_router.get("/maintenance", response_model=List[MaintenanceSchedule])
async def get_maintenance_schedules(asset_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        query["tenant_id"] = current_user.tenant_id
    
    if asset_id:
        query["asset_id"] = asset_id
    
    schedules = await db.maintenance_schedules.find(query, {"_id": 0}).to_list(1000)
    return [MaintenanceSchedule(**s) for s in schedules]

@api_router.patch("/maintenance/{maintenance_id}", response_model=MaintenanceSchedule)
async def update_maintenance(maintenance_id: str, maintenance_update: MaintenanceScheduleUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    maintenance = await db.maintenance_schedules.find_one({"id": maintenance_id}, {"_id": 0})
    if not maintenance:
        raise HTTPException(status_code=404, detail="Maintenance schedule not found")
    
    update_data = maintenance_update.model_dump(exclude_unset=True)
    await db.maintenance_schedules.update_one({"id": maintenance_id}, {"$set": update_data})
    
    updated = await db.maintenance_schedules.find_one({"id": maintenance_id}, {"_id": 0})
    return MaintenanceSchedule(**updated)

# Asset checkout/return endpoints
@api_router.post("/assets/{asset_id}/checkout", response_model=Asset)
async def checkout_asset(asset_id: str, checkout_data: AssetCheckout, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if asset["status"] not in [AssetStatus.AVAILABLE, AssetStatus.IN_USE]:
        raise HTTPException(status_code=400, detail="Asset not available for checkout")
    
    update_data = {
        "status": AssetStatus.CHECKED_OUT,
        "checked_out_to": checkout_data.checked_out_to,
        "checkout_date": datetime.now(timezone.utc).isoformat(),
        "expected_return_date": checkout_data.expected_return_date,
        "actual_return_date": None
    }
    
    await db.assets.update_one({"id": asset_id}, {"$set": update_data})
    
    history = AssetHistory(
        asset_id=asset_id,
        action="checked_out",
        performed_by=current_user.id,
        notes=f"Checked out to {checkout_data.checked_out_to}, expected return: {checkout_data.expected_return_date}"
    )
    await db.asset_history.insert_one(history.model_dump())
    
    updated_asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    return Asset(**updated_asset)

@api_router.post("/assets/{asset_id}/return", response_model=Asset)
async def return_asset(asset_id: str, return_data: AssetReturn, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if asset["status"] != AssetStatus.CHECKED_OUT:
        raise HTTPException(status_code=400, detail="Asset is not checked out")
    
    now_dt = datetime.now(timezone.utc)
    fine_amount = 0.0
    days_overdue = 0
    expected_return = asset.get("expected_return_date")
    if expected_return:
        try:
            due_dt = datetime.fromisoformat(expected_return.replace("Z", "+00:00"))
            if now_dt > due_dt:
                days_overdue = (now_dt - due_dt).days
                fine_amount = round(days_overdue * return_data.daily_fine_rate, 2)
        except Exception:
            pass

    update_data = {
        "status": AssetStatus.AVAILABLE,
        "actual_return_date": now_dt.isoformat(),
        "return_condition": return_data.return_condition,
        "last_fine_amount": fine_amount,
        "last_days_overdue": days_overdue,
    }

    await db.assets.update_one({"id": asset_id}, {"$set": update_data})

    notes_parts = [return_data.condition_notes or "Asset returned"]
    notes_parts.append(f"Condition: {return_data.return_condition}")
    if days_overdue > 0:
        notes_parts.append(f"Returned {days_overdue} day(s) late. Fine: ${fine_amount:.2f}")

    history = AssetHistory(
        asset_id=asset_id,
        action="returned",
        performed_by=current_user.id,
        notes=" | ".join(notes_parts)
    )
    await db.asset_history.insert_one(history.model_dump())

    updated_asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    result = Asset(**updated_asset).model_dump()
    result["days_overdue"] = days_overdue
    result["fine_amount"] = fine_amount
    return result

# Asset depreciation calculation
@api_router.get("/assets/{asset_id}/depreciation")
async def calculate_depreciation(asset_id: str, current_user: User = Depends(get_current_user)):
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if not asset.get("purchase_date") or not asset.get("purchase_price"):
        return {"current_value": asset.get("purchase_price", 0), "depreciation": 0, "years": 0}
    
    purchase_date = datetime.fromisoformat(asset["purchase_date"].replace('Z', '+00:00'))
    years_owned = (datetime.now(timezone.utc) - purchase_date).days / 365.25
    
    purchase_price = asset.get("purchase_price", 0)
    salvage_value = asset.get("salvage_value", 0)
    depreciation_rate = asset.get("depreciation_rate", 20) / 100
    method = asset.get("depreciation_method", "straight_line")
    
    if method == "straight_line":
        annual_depreciation = (purchase_price - salvage_value) * depreciation_rate
        total_depreciation = min(annual_depreciation * years_owned, purchase_price - salvage_value)
        current_value = max(purchase_price - total_depreciation, salvage_value)
    elif method == "declining_balance":
        current_value = purchase_price * ((1 - depreciation_rate) ** years_owned)
        current_value = max(current_value, salvage_value)
        total_depreciation = purchase_price - current_value
    else:
        current_value = purchase_price
        total_depreciation = 0
    
    # Update current value in database
    await db.assets.update_one({"id": asset_id}, {"$set": {"current_value": current_value}})
    
    return {
        "asset_id": asset_id,
        "purchase_price": purchase_price,
        "current_value": round(current_value, 2),
        "total_depreciation": round(total_depreciation, 2),
        "years_owned": round(years_owned, 2),
        "depreciation_method": method,
        "annual_depreciation_rate": asset.get("depreciation_rate", 20)
    }

# Warranty expiry check
@api_router.get("/assets/warranty/expiring")
async def get_expiring_warranties(days: int = 30, current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        query["tenant_id"] = current_user.tenant_id
    
    assets = await db.assets.find(query, {"_id": 0}).to_list(1000)
    
    expiring = []
    current_date = datetime.now(timezone.utc)
    threshold_date = current_date + timedelta(days=days)
    
    for asset in assets:
        if asset.get("warranty_end_date"):
            warranty_end = datetime.fromisoformat(asset["warranty_end_date"].replace('Z', '+00:00'))
            if current_date <= warranty_end <= threshold_date:
                days_remaining = (warranty_end - current_date).days
                expiring.append({
                    **asset,
                    "days_until_expiry": days_remaining
                })
    
    return expiring

# ===== Subscription Tier Endpoints =====

async def seed_default_tiers():
    """Seed default subscription tiers if none exist"""
    count = await db.subscription_tiers.count_documents({})
    if count > 0:
        return
    
    default_tiers = [
        {
            "id": "tier-free",
            "name": "Free",
            "slug": "free",
            "description": "Get started with basic IT asset management",
            "sort_order": 0,
            "is_default": True,
            "limits": {
                "max_users": 3,
                "max_assets": 10,
                "max_orders_per_month": 5,
                "max_tickets_per_month": 10,
            },
            "allowed_features": ["products", "orders", "assets", "tickets"],
            "highlights": ["Up to 3 users", "10 assets", "5 orders/month", "10 tickets/month", "Basic catalog"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": "tier-pro",
            "name": "Pro",
            "slug": "pro",
            "description": "For growing teams that need full IT lifecycle management",
            "sort_order": 1,
            "is_default": False,
            "limits": {
                "max_users": 25,
                "max_assets": 100,
                "max_orders_per_month": 50,
                "max_tickets_per_month": -1,
            },
            "allowed_features": ["products", "orders", "assets", "tickets", "users", "groups", "workflows"],
            "highlights": ["Up to 25 users", "100 assets", "50 orders/month", "Unlimited tickets", "Groups & Workflows", "White-label branding"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": "tier-enterprise",
            "name": "Enterprise",
            "slug": "enterprise",
            "description": "Full-featured platform for large organizations",
            "sort_order": 2,
            "is_default": False,
            "limits": {
                "max_users": -1,
                "max_assets": -1,
                "max_orders_per_month": -1,
                "max_tickets_per_month": -1,
            },
            "allowed_features": ["products", "orders", "assets", "tickets", "users", "groups", "workflows", "api_access", "priority_support"],
            "highlights": ["Unlimited users", "Unlimited assets", "Unlimited orders", "Unlimited tickets", "API access", "Priority support", "Custom integrations"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    await db.subscription_tiers.insert_many(default_tiers)
    
    # Assign default tier to existing tenants that don't have one
    await db.tenants.update_many(
        {"subscription_tier_id": {"$exists": False}},
        {"$set": {"subscription_tier_id": "tier-free", "subscription_started_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.tenants.update_many(
        {"subscription_tier_id": None},
        {"$set": {"subscription_tier_id": "tier-free", "subscription_started_at": datetime.now(timezone.utc).isoformat()}}
    )

@api_router.get("/subscription-tiers")
async def get_subscription_tiers(current_user: User = Depends(get_current_user)):
    tiers = await db.subscription_tiers.find({}, {"_id": 0}).to_list(100)
    tiers.sort(key=lambda t: t.get("sort_order", 0))
    return tiers

@api_router.get("/subscription-tiers/{tier_id}")
async def get_subscription_tier(tier_id: str, current_user: User = Depends(get_current_user)):
    tier = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")
    return tier

@api_router.post("/subscription-tiers")
async def create_subscription_tier(tier_data: SubscriptionTierCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can manage tiers")
    
    existing = await db.subscription_tiers.find_one({"slug": tier_data.slug}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Tier with this slug already exists")
    
    tier = SubscriptionTier(**tier_data.model_dump())
    await db.subscription_tiers.insert_one(tier.model_dump())
    return tier.model_dump()

@api_router.patch("/subscription-tiers/{tier_id}")
async def update_subscription_tier(tier_id: str, tier_update: SubscriptionTierUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can manage tiers")
    
    tier = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")
    
    update_data = tier_update.model_dump(exclude_unset=True)
    if update_data:
        await db.subscription_tiers.update_one({"id": tier_id}, {"$set": update_data})
    
    updated = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    return updated

@api_router.delete("/subscription-tiers/{tier_id}")
async def delete_subscription_tier(tier_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can manage tiers")
    
    tier = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")
    
    # Check if any tenants are using this tier
    tenant_count = await db.tenants.count_documents({"subscription_tier_id": tier_id})
    if tenant_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete tier: {tenant_count} tenant(s) are using it")
    
    await db.subscription_tiers.delete_one({"id": tier_id})
    return {"message": "Tier deleted"}

@api_router.patch("/tenants/{tenant_id}/subscription")
async def update_tenant_subscription(tenant_id: str, tier_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can change subscriptions")
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    tier = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(status_code=404, detail="Tier not found")
    
    # Update tenant subscription and sync enabled features
    await db.tenants.update_one(
        {"id": tenant_id},
        {"$set": {
            "subscription_tier_id": tier_id,
            "subscription_started_at": datetime.now(timezone.utc).isoformat(),
            "enabled_features": tier["allowed_features"]
        }}
    )
    
    updated = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    return Tenant(**updated)

@api_router.get("/tenants/{tenant_id}/usage")
async def get_tenant_usage(tenant_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role == UserRole.EMPLOYEE:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role not in [UserRole.SUPER_ADMIN] and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    # Get current tier
    tier_id = tenant.get("subscription_tier_id", "tier-free")
    tier = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        # Fallback to default tier
        tier = await db.subscription_tiers.find_one({"is_default": True}, {"_id": 0})
    
    # Calculate current usage
    user_count = await db.users.count_documents({"tenant_id": tenant_id})
    asset_count = await db.assets.count_documents({"tenant_id": tenant_id})
    
    # Monthly counts
    month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    order_count = await db.orders.count_documents({
        "tenant_id": tenant_id,
        "created_at": {"$gte": month_start.isoformat()}
    })
    ticket_count = await db.tickets.count_documents({
        "tenant_id": tenant_id,
        "created_at": {"$gte": month_start.isoformat()}
    })
    
    limits = tier.get("limits", {})
    
    usage = {
        "tenant_id": tenant_id,
        "tier": tier,
        "current_usage": {
            "users": user_count,
            "assets": asset_count,
            "orders_this_month": order_count,
            "tickets_this_month": ticket_count,
        },
        "limits": {
            "max_users": limits.get("max_users", 3),
            "max_assets": limits.get("max_assets", 10),
            "max_orders_per_month": limits.get("max_orders_per_month", 5),
            "max_tickets_per_month": limits.get("max_tickets_per_month", 10),
        },
        "subscription_started_at": tenant.get("subscription_started_at"),
    }
    
    return usage

# ── Billing / Invoices ────────────────────────────────────────────────────────

@api_router.get("/invoices")
async def list_invoices(current_user: User = Depends(get_current_user)):
    if current_user.role == UserRole.EMPLOYEE:
        raise HTTPException(status_code=403, detail="Access denied")
    query = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return invoices

@api_router.post("/invoices")
async def create_invoice(invoice_data: InvoiceCreate, tenant_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can create invoices")
    invoice = Invoice(tenant_id=tenant_id, **invoice_data.model_dump())
    await db.invoices.insert_one(invoice.model_dump())
    return invoice

@api_router.patch("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, update: InvoiceUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admins can update invoices")
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    return await db.invoices.find_one({"id": invoice_id}, {"_id": 0})

# ── DELETE endpoints ──────────────────────────────────────────────────────────

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    # Soft delete
    await db.products.update_one({"id": product_id}, {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat()}})
    await write_audit_log(current_user.id, "delete", "product", product_id)
    return {"message": "Product deleted successfully"}

@api_router.delete("/assets/{asset_id}")
async def delete_asset(asset_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if current_user.role != UserRole.SUPER_ADMIN and asset.get("tenant_id") != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    # Soft delete
    await db.assets.update_one({"id": asset_id}, {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat()}})
    await write_audit_log(current_user.id, "delete", "asset", asset_id)
    return {"message": "Asset deleted successfully"}

@api_router.delete("/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        raise HTTPException(status_code=403, detail="Access denied")
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    await db.tickets.delete_one({"id": ticket_id})
    await write_audit_log(current_user.id, "delete", "ticket", ticket_id)
    return {"message": "Ticket deleted successfully"}

@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") == OrderStatus.FULFILLED:
        raise HTTPException(status_code=400, detail="Cannot delete a fulfilled order")
    await db.orders.delete_one({"id": order_id})
    await write_audit_log(current_user.id, "delete", "order", order_id)
    return {"message": "Order deleted successfully"}

# ── Feature 1: Notifications endpoint ────────────────────────────────────────

@api_router.get("/notifications")
async def get_notifications(current_user: User = Depends(get_current_user)):
    notifications = []
    tenant_id = current_user.tenant_id

    # Pending orders awaiting approval (for admins/managers)
    if current_user.role in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        query = {"status": "pending"}
        if tenant_id:
            query["tenant_id"] = tenant_id
        pending_count = await db.orders.count_documents(query)
        if pending_count > 0:
            notifications.append({
                "type": "pending_order",
                "message": f"{pending_count} order{'s' if pending_count > 1 else ''} awaiting approval",
                "count": pending_count,
                "link": "/orders"
            })

    # Expiring warranties (within 30 days)
    if current_user.role in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        asset_query = {"is_deleted": {"$ne": True}}
        if tenant_id:
            asset_query["tenant_id"] = tenant_id
        assets = await db.assets.find(asset_query, {"_id": 0, "warranty_end_date": 1}).to_list(1000)
        now = datetime.now(timezone.utc)
        threshold = now + timedelta(days=30)
        expiring = sum(
            1 for a in assets
            if a.get("warranty_end_date") and
            now <= datetime.fromisoformat(a["warranty_end_date"].replace("Z", "+00:00")) <= threshold
        )
        if expiring > 0:
            notifications.append({
                "type": "expiring_warranty",
                "message": f"{expiring} warranty{'s' if expiring > 1 else ''} expiring within 30 days",
                "count": expiring,
                "link": "/assets"
            })

    # Overdue maintenance (correct collection)
    maint_query = {"status": "overdue"}
    if tenant_id:
        maint_query["tenant_id"] = tenant_id
    overdue = await db.maintenance_schedules.count_documents(maint_query)
    if overdue > 0:
        notifications.append({
            "type": "overdue_maintenance",
            "message": f"{overdue} maintenance task{'s' if overdue > 1 else ''} overdue",
            "count": overdue,
            "link": "/assets"
        })

    # Open tickets assigned to me
    if current_user.role in [UserRole.HELPDESK_AGENT, UserRole.ASSET_MANAGER]:
        my_tickets = await db.tickets.count_documents({
            "assigned_to": current_user.id,
            "status": {"$in": ["open", "in_progress"]}
        })
        if my_tickets > 0:
            notifications.append({
                "type": "assigned_ticket",
                "message": f"{my_tickets} ticket{'s' if my_tickets > 1 else ''} assigned to you",
                "count": my_tickets,
                "link": "/tickets"
            })

    return notifications

# ── Feature 1: Email alerts trigger ──────────────────────────────────────────

@api_router.post("/alerts/trigger")
async def trigger_email_alerts(current_user: User = Depends(get_current_user)):
    """Manually trigger alert emails to all tenant admins."""
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")

    tenant_id = current_user.tenant_id
    sent_count = 0

    # Find admin emails to notify
    admin_query = {"role": {"$in": ["tenant_admin", "super_admin"]}}
    if tenant_id:
        admin_query["tenant_id"] = tenant_id
    admins = await db.users.find(admin_query, {"_id": 0, "email": 1, "name": 1}).to_list(100)
    admin_emails = [a["email"] for a in admins if a.get("email")]

    # Check expiring warranties
    asset_query = {"is_deleted": {"$ne": True}}
    if tenant_id:
        asset_query["tenant_id"] = tenant_id
    assets = await db.assets.find(asset_query, {"_id": 0}).to_list(1000)
    now = datetime.now(timezone.utc)
    threshold = now + timedelta(days=30)
    expiring_assets = [
        f"{a.get('asset_tag', 'Unknown')} — warranty ends {a['warranty_end_date'][:10]}"
        for a in assets
        if a.get("warranty_end_date") and
        now <= datetime.fromisoformat(a["warranty_end_date"].replace("Z", "+00:00")) <= threshold
    ]
    if expiring_assets:
        for email in admin_emails:
            await send_alert_email(
                email,
                f"⚠️ {len(expiring_assets)} Warranty Expiring Soon",
                expiring_assets,
                "The following assets have warranties expiring within 30 days:"
            )
        sent_count += len(admin_emails)

    # Check overdue maintenance (use correct collection name)
    maint_query = {"status": {"$in": ["scheduled", "overdue", "in_progress"]}}
    if tenant_id:
        maint_query["tenant_id"] = tenant_id
    schedules = await db.maintenance_schedules.find(maint_query, {"_id": 0}).to_list(500)
    now_date = now.date().isoformat()
    overdue_maint = [
        f"{s.get('title', 'Untitled')} — was due {s['scheduled_date'][:10]}"
        for s in schedules
        if s.get("scheduled_date") and s["scheduled_date"][:10] < now_date
    ]
    if overdue_maint:
        for email in admin_emails:
            await send_alert_email(
                email,
                f"🔧 {len(overdue_maint)} Maintenance Tasks Overdue",
                overdue_maint,
                "The following maintenance tasks are past their scheduled date:"
            )
        sent_count += len(admin_emails)

    # Check pending orders
    order_query = {"status": "pending"}
    if tenant_id:
        order_query["tenant_id"] = tenant_id
    pending_orders = await db.orders.find(order_query, {"_id": 0}).to_list(100)
    if pending_orders:
        items = [f"Order ID: {o['id'][:8]}... — Qty: {o.get('quantity', 1)}" for o in pending_orders]
        for email in admin_emails:
            await send_alert_email(
                email,
                f"📦 {len(pending_orders)} Orders Awaiting Approval",
                items,
                "The following orders are waiting for your approval:"
            )
        sent_count += len(admin_emails)

    await write_audit_log(current_user.id, "trigger_alerts", "system", "alerts", f"Sent to {sent_count} recipients")
    return {"message": f"Alert emails sent to {len(admin_emails)} admin(s)", "alerts_triggered": sent_count}

# ── Feature 3: Asset History endpoint ────────────────────────────────────────

@api_router.get("/assets/{asset_id}/history")
async def get_asset_history(asset_id: str, current_user: User = Depends(get_current_user)):
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    history = await db.asset_history.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("date", -1).to_list(200)

    # Enrich with user names
    user_ids = list({h["performed_by"] for h in history if h.get("performed_by")})
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(100)
    user_map = {u["id"]: u for u in users}

    for h in history:
        uid = h.get("performed_by")
        h["performed_by_name"] = user_map.get(uid, {}).get("name", "Unknown")

    return history

# ── Feature 5: Department endpoints ──────────────────────────────────────────

@api_router.post("/departments", response_model=Department)
async def create_department(dept_data: DepartmentCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role == UserRole.TENANT_ADMIN and dept_data.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Cannot create departments for other tenants")
    dept = Department(**dept_data.model_dump())
    await db.departments.insert_one(dept.model_dump())
    return dept

@api_router.get("/departments", response_model=List[Department])
async def get_departments(current_user: User = Depends(get_current_user)):
    query: dict = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id
    depts = await db.departments.find(query, {"_id": 0}).to_list(500)
    return [Department(**d) for d in depts]

@api_router.patch("/departments/{dept_id}", response_model=Department)
async def update_department(dept_id: str, update: DepartmentUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    dept = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    if current_user.role == UserRole.TENANT_ADMIN and dept["tenant_id"] != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    update_data = update.model_dump(exclude_unset=True)
    await db.departments.update_one({"id": dept_id}, {"$set": update_data})
    updated = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    return Department(**updated)

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    dept = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    await db.departments.delete_one({"id": dept_id})
    return {"message": "Department deleted"}

@api_router.get("/departments/{dept_id}/stats")
async def get_department_stats(dept_id: str, current_user: User = Depends(get_current_user)):
    dept = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    asset_count = await db.assets.count_documents({"department_id": dept_id, "is_deleted": {"$ne": True}})
    user_count = await db.users.count_documents({"department_id": dept_id})
    assets = await db.assets.find({"department_id": dept_id, "is_deleted": {"$ne": True}}, {"_id": 0, "purchase_price": 1}).to_list(5000)
    total_value = sum(a.get("purchase_price", 0) for a in assets)
    return {
        "department_id": dept_id,
        "name": dept["name"],
        "asset_count": asset_count,
        "user_count": user_count,
        "total_asset_value": round(total_value, 2),
        "budget": dept.get("budget", 0)
    }

# ── Bulk operations ───────────────────────────────────────────────────────────

class BulkDeleteRequest(BaseModel):
    ids: List[str]

@api_router.post("/assets/bulk-delete")
async def bulk_delete_assets(body: BulkDeleteRequest, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    if not body.ids:
        raise HTTPException(status_code=400, detail="No asset IDs provided")

    query: dict = {"id": {"$in": body.ids}}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    result = await db.assets.update_many(
        query,
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat()}}
    )
    await write_audit_log(current_user.id, "bulk_delete", "assets", ",".join(body.ids), f"{result.modified_count} assets deleted")
    return {"message": f"{result.modified_count} assets deleted successfully"}

@api_router.post("/tickets/bulk-delete")
async def bulk_delete_tickets(body: BulkDeleteRequest, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER, UserRole.HELPDESK_AGENT]:
        raise HTTPException(status_code=403, detail="Access denied")
    if not body.ids:
        raise HTTPException(status_code=400, detail="No ticket IDs provided")

    query: dict = {"id": {"$in": body.ids}}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    result = await db.tickets.delete_many(query)
    return {"message": f"{result.deleted_count} tickets deleted successfully"}

# ── CSV Export endpoints ───────────────────────────────────────────────────────

@api_router.get("/assets/export/csv")
async def export_assets_csv(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    query: dict = {"is_deleted": {"$ne": True}}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    assets = await db.assets.find(query, {"_id": 0}).to_list(10000)

    import io, csv as csv_module
    output = io.StringIO()
    fieldnames = ["id", "name", "asset_tag", "serial_number", "category", "status",
                  "assigned_to", "purchase_price", "purchase_date", "warranty_end_date", "location"]
    writer = csv_module.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for asset in assets:
        writer.writerow(asset)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets_export.csv"}
    )

@api_router.get("/orders/export/csv")
async def export_orders_csv(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    query: dict = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    orders = await db.orders.find(query, {"_id": 0}).to_list(10000)

    import io, csv as csv_module
    output = io.StringIO()
    fieldnames = ["id", "user_id", "product_id", "quantity", "status", "approved_by", "approval_date", "created_at"]
    writer = csv_module.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for order in orders:
        writer.writerow(order)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders_export.csv"}
    )

# ── Audit log endpoint ────────────────────────────────────────────────────────

@api_router.get("/audit-log")
async def get_audit_log(
    current_user: User = Depends(get_current_user),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200)
):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    skip = (page - 1) * limit
    logs = await db.audit_log.find({}, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return logs

# ═══════════════════════════════════════════════════════════════════════════
# USER PROFILE & PASSWORD CHANGE
# ═══════════════════════════════════════════════════════════════════════════

@api_router.get("/profile")
async def get_profile(current_user: User = Depends(get_current_user_hybrid)):
    """Return current user profile with assigned assets, open tickets, recent orders."""
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0, "password_hash": 0})
    assets = await db.assets.find({"assigned_to": current_user.id, "is_deleted": {"$ne": True}}, {"_id": 0}).to_list(100)
    tickets = await db.tickets.find({"created_by": current_user.id, "status": {"$in": ["open", "in_progress"]}}, {"_id": 0}).sort("created_at", -1).to_list(20)
    orders = await db.orders.find({"user_id": current_user.id}, {"_id": 0}).sort("created_at", -1).to_list(20)
    return {
        "user": user_doc,
        "assigned_assets": assets,
        "open_tickets": tickets,
        "recent_orders": orders,
    }

@api_router.put("/profile")
async def update_profile(update: ProfileUpdate, current_user: User = Depends(get_current_user_hybrid)):
    updates = {k: v for k, v in update.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.users.update_one({"id": current_user.id}, {"$set": updates})
    return {"message": "Profile updated"}

@api_router.put("/auth/change-password")
async def change_password(req: ChangePasswordRequest, current_user: User = Depends(get_current_user_hybrid)):
    user_doc = await db.users.find_one({"id": current_user.id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    if not verify_password(req.current_password, user_doc.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    validate_password_strength(req.new_password)
    new_hash = hash_password(req.new_password)
    await db.users.update_one({"id": current_user.id}, {"$set": {"password_hash": new_hash}})
    await write_audit_log(current_user.id, "change_password", "user", current_user.id)
    return {"message": "Password changed successfully"}

# ═══════════════════════════════════════════════════════════════════════════
# LOCATION MASTER
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/locations")
async def create_location(loc: LocationCreate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    location = Location(**loc.model_dump())
    await db.locations.insert_one(location.model_dump())
    return location

@api_router.get("/locations")
async def list_locations(current_user: User = Depends(get_current_user_hybrid)):
    query = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    locs = await db.locations.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return locs

@api_router.put("/locations/{location_id}")
async def update_location(location_id: str, update: LocationUpdate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    updates = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.locations.update_one({"id": location_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    return {"message": "Location updated"}

@api_router.delete("/locations/{location_id}")
async def delete_location(location_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.locations.delete_one({"id": location_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    return {"message": "Location deleted"}

# ═══════════════════════════════════════════════════════════════════════════
# VENDOR MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/vendors")
async def create_vendor(vendor: VendorCreate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    v = Vendor(**vendor.model_dump())
    await db.vendors.insert_one(v.model_dump())
    return v

@api_router.get("/vendors")
async def list_vendors(current_user: User = Depends(get_current_user_hybrid)):
    query = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    vendors = await db.vendors.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return vendors

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, update: VendorUpdate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    updates = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.vendors.update_one({"id": vendor_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor updated"}

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.vendors.delete_one({"id": vendor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor deleted"}

# ═══════════════════════════════════════════════════════════════════════════
# SOFTWARE LICENSE TRACKING
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/licenses")
async def create_license(lic: LicenseCreate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    license_obj = SoftwareLicense(**lic.model_dump())
    await db.licenses.insert_one(license_obj.model_dump())
    return license_obj

@api_router.get("/licenses")
async def list_licenses(current_user: User = Depends(get_current_user_hybrid)):
    query = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    licenses = await db.licenses.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return licenses

@api_router.put("/licenses/{license_id}")
async def update_license(license_id: str, update: LicenseUpdate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    updates = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.licenses.update_one({"id": license_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="License not found")
    return {"message": "License updated"}

@api_router.post("/licenses/seed-demo")
async def seed_demo_licenses(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")

    tenant_id = current_user.tenant_id
    now = datetime.now(timezone.utc)

    demo_licenses = [
        {"name": "Microsoft 365", "vendor": "Microsoft", "license_type": "subscription",
         "seats_total": 50, "seats_used": 42, "cost": 85000,
         "purchase_date": "2024-01-01", "expiry_date": "2025-12-31",
         "license_key": "M365-DEMO-XXXX-XXXX", "notes": "Company-wide Office suite"},
        {"name": "Adobe Creative Cloud", "vendor": "Adobe", "license_type": "subscription",
         "seats_total": 20, "seats_used": 12, "cost": 48000,
         "purchase_date": "2024-03-01", "expiry_date": "2025-09-30",
         "license_key": "ACC-DEMO-XXXX-XXXX", "notes": "Design team licenses"},
        {"name": "Zoom Business", "vendor": "Zoom", "license_type": "subscription",
         "seats_total": 30, "seats_used": 28, "cost": 32000,
         "purchase_date": "2024-01-15", "expiry_date": "2025-11-30",
         "license_key": "ZOOM-DEMO-XXXX-XXXX", "notes": "Video conferencing"},
        {"name": "Slack Pro", "vendor": "Slack", "license_type": "subscription",
         "seats_total": 40, "seats_used": 35, "cost": 28000,
         "purchase_date": "2024-02-01", "expiry_date": "2026-01-31",
         "license_key": "SLACK-DEMO-XXXX-XXXX", "notes": "Team communication"},
        {"name": "GitHub Enterprise", "vendor": "GitHub", "license_type": "subscription",
         "seats_total": 15, "seats_used": 11, "cost": 55000,
         "purchase_date": "2024-04-01", "expiry_date": "2025-03-31",
         "license_key": "GHE-DEMO-XXXX-XXXX", "notes": "Developer team"},
        {"name": "Antivirus Pro", "vendor": "Kaspersky", "license_type": "subscription",
         "seats_total": 100, "seats_used": 87, "cost": 18000,
         "purchase_date": "2024-01-01", "expiry_date": "2025-07-31",
         "license_key": "KAV-DEMO-XXXX-XXXX", "notes": "Endpoint protection"},
        {"name": "AutoCAD", "vendor": "Autodesk", "license_type": "perpetual",
         "seats_total": 5, "seats_used": 3, "cost": 120000,
         "purchase_date": "2023-06-01", "expiry_date": None,
         "license_key": "ACAD-DEMO-XXXX-XXXX", "notes": "Engineering team"},
    ]

    inserted = 0
    for d in demo_licenses:
        existing = await db.licenses.find_one(
            {"name": {"$regex": f"^{re.escape(d['name'])}$", "$options": "i"}, "tenant_id": tenant_id}
        )
        if not existing:
            lic = SoftwareLicense(
                tenant_id=tenant_id,
                name=d["name"], vendor=d["vendor"], license_type=d["license_type"],
                seats_total=d["seats_total"], seats_used=d["seats_used"], cost=d["cost"],
                purchase_date=d["purchase_date"], expiry_date=d["expiry_date"],
                license_key=d["license_key"], notes=d["notes"],
            )
            await db.licenses.insert_one(lic.model_dump())
            inserted += 1

    return {"message": f"{inserted} demo licenses added", "inserted": inserted}

@api_router.delete("/licenses/{license_id}")
async def delete_license(license_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.licenses.delete_one({"id": license_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="License not found")
    return {"message": "License deleted"}

# ═══════════════════════════════════════════════════════════════════════════
# API KEYS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/auth/api-keys", response_model=ApiKey)
async def create_api_key(data: ApiKeyCreate, current_user: User = Depends(get_current_user)):
    """Generate a new API key for the current user."""
    key_obj = ApiKey(user_id=current_user.id, tenant_id=current_user.tenant_id, name=data.name)
    await db.api_keys.insert_one(key_obj.model_dump())
    return key_obj

@api_router.get("/auth/api-keys")
async def list_api_keys(current_user: User = Depends(get_current_user)):
    keys = await db.api_keys.find({"user_id": current_user.id, "is_active": True}, {"_id": 0}).to_list(50)
    # Mask the key to show only prefix + last 4 chars
    result = []
    for k in keys:
        masked = k["key"][:8] + "..." + k["key"][-4:]
        result.append({**k, "key_display": masked})
    return result

@api_router.delete("/auth/api-keys/{key_id}")
async def revoke_api_key(key_id: str, current_user: User = Depends(get_current_user)):
    key = await db.api_keys.find_one({"id": key_id, "user_id": current_user.id})
    if not key:
        raise HTTPException(status_code=404, detail="API key not found")
    await db.api_keys.update_one({"id": key_id}, {"$set": {"is_active": False}})
    return {"message": "API key revoked"}

# ═══════════════════════════════════════════════════════════════════════════
# ASSET RESERVATIONS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/reservations")
async def create_reservation(res: ReservationCreate, current_user: User = Depends(get_current_user_hybrid)):
    asset = await db.assets.find_one({"id": res.asset_id, "is_deleted": {"$ne": True}})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if asset.get("status") not in ["available"]:
        raise HTTPException(status_code=400, detail="Asset is not available for reservation")
    reservation = AssetReservation(
        **res.model_dump(),
        reserved_by=current_user.id,
        tenant_id=current_user.tenant_id or asset["tenant_id"],
    )
    await db.reservations.insert_one(reservation.model_dump())
    return reservation

@api_router.get("/reservations")
async def list_reservations(current_user: User = Depends(get_current_user_hybrid)):
    query = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    if current_user.role == UserRole.EMPLOYEE:
        query["reserved_by"] = current_user.id
    reservations = await db.reservations.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return reservations

@api_router.put("/reservations/{reservation_id}")
async def update_reservation(reservation_id: str, status: str, current_user: User = Depends(get_current_user_hybrid)):
    """Approve, reject, or cancel a reservation. status: approved|rejected|cancelled"""
    valid = {"approved", "rejected", "cancelled", "completed"}
    if status not in valid:
        raise HTTPException(status_code=400, detail=f"status must be one of {valid}")
    res = await db.reservations.find_one({"id": reservation_id})
    if not res:
        raise HTTPException(status_code=404, detail="Reservation not found")
    if status == "cancelled" and res["reserved_by"] != current_user.id:
        if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
            raise HTTPException(status_code=403, detail="Access denied")
    elif status in {"approved", "rejected"} and current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    await db.reservations.update_one({"id": reservation_id}, {"$set": {"status": status}})
    return {"message": f"Reservation {status}"}

@api_router.delete("/reservations/{reservation_id}")
async def delete_reservation(reservation_id: str, current_user: User = Depends(get_current_user_hybrid)):
    res = await db.reservations.find_one({"id": reservation_id})
    if not res:
        raise HTTPException(status_code=404, detail="Reservation not found")
    if res["reserved_by"] != current_user.id and current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    await db.reservations.delete_one({"id": reservation_id})
    return {"message": "Reservation deleted"}

# ═══════════════════════════════════════════════════════════════════════════
# TICKET COMMENTS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/tickets/{ticket_id}/comments")
async def add_comment(ticket_id: str, body: TicketCommentCreate, current_user: User = Depends(get_current_user_hybrid)):
    ticket = await db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    comment = TicketComment(
        ticket_id=ticket_id,
        author_id=current_user.id,
        author_name=current_user.name,
        content=body.content,
    )
    await db.ticket_comments.insert_one(comment.model_dump())
    await db.tickets.update_one({"id": ticket_id}, {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}})
    return comment

@api_router.get("/tickets/{ticket_id}/comments")
async def get_comments(ticket_id: str, current_user: User = Depends(get_current_user_hybrid)):
    comments = await db.ticket_comments.find({"ticket_id": ticket_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return comments

@api_router.delete("/tickets/{ticket_id}/comments/{comment_id}")
async def delete_comment(ticket_id: str, comment_id: str, current_user: User = Depends(get_current_user_hybrid)):
    comment = await db.ticket_comments.find_one({"id": comment_id, "ticket_id": ticket_id})
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")
    if comment["author_id"] != current_user.id and current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    await db.ticket_comments.delete_one({"id": comment_id})
    return {"message": "Comment deleted"}

# ═══════════════════════════════════════════════════════════════════════════
# ASSET TRANSFER REQUESTS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/transfers")
async def create_transfer(req: TransferCreate, current_user: User = Depends(get_current_user_hybrid)):
    asset = await db.assets.find_one({"id": req.asset_id, "is_deleted": {"$ne": True}})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    transfer = AssetTransfer(
        asset_id=req.asset_id,
        asset_tag=asset.get("asset_tag", ""),
        from_user_id=asset.get("assigned_to", current_user.id),
        to_user_id=req.to_user_id,
        requested_by=current_user.id,
        tenant_id=current_user.tenant_id or asset["tenant_id"],
        reason=req.reason,
    )
    await db.asset_transfers.insert_one(transfer.model_dump())
    return transfer

@api_router.get("/transfers")
async def list_transfers(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role == UserRole.SUPER_ADMIN:
        query = {}
    elif current_user.role in [UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        query = {"tenant_id": current_user.tenant_id}
    else:
        query = {"$or": [{"from_user_id": current_user.id}, {"to_user_id": current_user.id}, {"requested_by": current_user.id}]}
    transfers = await db.asset_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return transfers

@api_router.put("/transfers/{transfer_id}")
async def update_transfer(transfer_id: str, action: str, current_user: User = Depends(get_current_user_hybrid)):
    """action: approved|rejected|completed|cancelled"""
    valid = {"approved", "rejected", "completed", "cancelled"}
    if action not in valid:
        raise HTTPException(status_code=400, detail=f"action must be one of {valid}")
    transfer = await db.asset_transfers.find_one({"id": transfer_id})
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    updates: dict = {"status": action}
    if action == "approved":
        if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
            raise HTTPException(status_code=403, detail="Access denied")
        updates["approved_by"] = current_user.id
        updates["approved_at"] = datetime.now(timezone.utc).isoformat()
    if action == "completed":
        # Actually move the asset
        await db.assets.update_one(
            {"id": transfer["asset_id"]},
            {"$set": {"assigned_to": transfer["to_user_id"], "status": "assigned"}}
        )
        await write_audit_log(current_user.id, "transfer_complete", "asset", transfer["asset_id"],
                              f"Transferred to {transfer['to_user_id']}")
    await db.asset_transfers.update_one({"id": transfer_id}, {"$set": updates})
    return {"message": f"Transfer {action}"}

# ═══════════════════════════════════════════════════════════════════════════
# CUSTOM FIELDS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/custom-fields")
async def create_custom_field(field: CustomFieldCreate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    cf = CustomField(**field.model_dump())
    await db.custom_fields.insert_one(cf.model_dump())
    return cf

@api_router.get("/custom-fields")
async def list_custom_fields(current_user: User = Depends(get_current_user_hybrid)):
    query = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    fields = await db.custom_fields.find(query, {"_id": 0}).sort("field_label", 1).to_list(200)
    return fields

@api_router.delete("/custom-fields/{field_id}")
async def delete_custom_field(field_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.custom_fields.delete_one({"id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Custom field not found")
    return {"message": "Custom field deleted"}

@api_router.put("/assets/{asset_id}/custom-values")
async def set_custom_values(asset_id: str, values: dict, current_user: User = Depends(get_current_user_hybrid)):
    result = await db.assets.update_one({"id": asset_id}, {"$set": {"custom_values": values}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {"message": "Custom values updated"}

# ═══════════════════════════════════════════════════════════════════════════
# BULK EDIT ASSETS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.put("/assets/bulk-update")
async def bulk_update_assets(body: BulkAssetUpdate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    if not body.ids:
        raise HTTPException(status_code=400, detail="No asset IDs provided")
    updates = {}
    if body.status is not None:
        updates["status"] = body.status.value
    if body.location is not None:
        updates["location"] = body.location
    if body.department_id is not None:
        updates["department_id"] = body.department_id
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    query = {"id": {"$in": body.ids}}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id
    result = await db.assets.update_many(query, {"$set": updates})
    await write_audit_log(current_user.id, "bulk_update", "assets", ",".join(body.ids), f"{result.modified_count} assets updated")
    return {"updated": result.modified_count}

# ═══════════════════════════════════════════════════════════════════════════
# ASSET PHOTO UPLOAD
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/assets/{asset_id}/photo")
async def upload_asset_photo(asset_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    if file.content_type not in ["image/jpeg", "image/png", "image/webp", "image/gif"]:
        raise HTTPException(status_code=400, detail="Only JPEG, PNG, WEBP, GIF images allowed")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image must be under 5MB")
    import base64
    b64 = base64.b64encode(content).decode()
    photo_data_url = f"data:{file.content_type};base64,{b64}"
    result = await db.assets.update_one({"id": asset_id}, {"$set": {"photo_url": photo_data_url}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {"message": "Photo uploaded", "photo_url": photo_data_url[:100] + "..."}

@api_router.delete("/assets/{asset_id}/photo")
async def delete_asset_photo(asset_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.assets.update_one({"id": asset_id}, {"$unset": {"photo_url": ""}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {"message": "Photo removed"}

# ═══════════════════════════════════════════════════════════════════════════
# DASHBOARD CHARTS DATA
# ═══════════════════════════════════════════════════════════════════════════

@api_router.get("/dashboard/charts")
async def get_dashboard_charts(current_user: User = Depends(get_current_user_hybrid)):
    tenant_q = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}

    # Asset status breakdown
    statuses = ["available", "assigned", "in_use", "under_maintenance", "disposed", "checked_out"]
    asset_status_data = []
    for s in statuses:
        count = await db.assets.count_documents({**tenant_q, "status": s, "is_deleted": {"$ne": True}})
        if count > 0:
            asset_status_data.append({"name": s.replace("_", " ").title(), "value": count})

    # Monthly orders (last 6 months)
    monthly_orders = []
    now = datetime.now(timezone.utc)
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i * 30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 0:
            month_end = now
        else:
            next_m = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_m.replace(day=1)
        count = await db.orders.count_documents({**tenant_q, "created_at": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()}})
        monthly_orders.append({"month": month_start.strftime("%b %Y"), "orders": count})

    # Ticket trends (last 6 months)
    ticket_trends = []
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i * 30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 0:
            month_end = now
        else:
            next_m = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_m.replace(day=1)
        count = await db.tickets.count_documents({**tenant_q, "created_at": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()}})
        ticket_trends.append({"month": month_start.strftime("%b %Y"), "tickets": count})

    # Ticket by status
    ticket_statuses = ["open", "in_progress", "resolved", "closed"]
    ticket_status_data = []
    for s in ticket_statuses:
        count = await db.tickets.count_documents({**tenant_q, "status": s})
        if count > 0:
            ticket_status_data.append({"name": s.replace("_", " ").title(), "value": count})

    # License seat utilization
    licenses_raw = await db.licenses.find(tenant_q, {"_id": 0}).to_list(100)
    license_utilization = []
    for lic in licenses_raw:
        total = lic.get("seats_total", 0)
        used = lic.get("seats_used", 0)
        available = max(0, total - used)
        license_utilization.append({
            "name": lic.get("name", "Unknown"),
            "used": used,
            "available": available,
            "total": total,
        })

    return {
        "asset_status": asset_status_data,
        "monthly_orders": monthly_orders,
        "ticket_trends": ticket_trends,
        "ticket_status": ticket_status_data,
        "license_utilization": license_utilization,
    }

# ═══════════════════════════════════════════════════════════════════════════
# ADVANCED REPORTS
# ═══════════════════════════════════════════════════════════════════════════

@api_router.get("/reports/assets-by-department")
async def report_assets_by_department(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    tenant_q = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    departments = await db.departments.find(tenant_q, {"_id": 0}).to_list(200)
    result = []
    for dept in departments:
        count = await db.assets.count_documents({"department_id": dept["id"], "is_deleted": {"$ne": True}})
        total_value = 0.0
        assets = await db.assets.find({"department_id": dept["id"], "is_deleted": {"$ne": True}}, {"purchase_price": 1}).to_list(10000)
        total_value = sum(a.get("purchase_price", 0) for a in assets)
        result.append({"department": dept["name"], "asset_count": count, "total_value": total_value})
    # Add unassigned
    no_dept = await db.assets.count_documents({**tenant_q, "department_id": None, "is_deleted": {"$ne": True}})
    if no_dept > 0:
        result.append({"department": "No Department", "asset_count": no_dept, "total_value": 0})
    return result

@api_router.get("/reports/maintenance-costs")
async def report_maintenance_costs(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    tenant_q = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    now = datetime.now(timezone.utc)
    monthly_costs = []
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i * 30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 0:
            month_end = now
        else:
            next_m = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_m.replace(day=1)
        records = await db.maintenance_schedules.find(
            {**tenant_q, "completed_date": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()}},
            {"cost": 1}
        ).to_list(10000)
        total = sum((r.get("cost") or 0) for r in records)
        monthly_costs.append({"month": month_start.strftime("%b %Y"), "cost": round(total, 2)})
    return monthly_costs

@api_router.get("/reports/expiring-warranties")
async def report_expiring_warranties(
    days: int = Query(90, ge=1, le=365),
    current_user: User = Depends(get_current_user_hybrid)
):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    tenant_q = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    cutoff = (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()
    today = datetime.now(timezone.utc).isoformat()
    assets = await db.assets.find(
        {**tenant_q, "warranty_end_date": {"$gte": today, "$lte": cutoff}, "is_deleted": {"$ne": True}},
        {"_id": 0}
    ).sort("warranty_end_date", 1).to_list(500)
    return assets

@api_router.get("/reports/asset-depreciation")
async def report_asset_depreciation(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    tenant_q = {} if current_user.role == UserRole.SUPER_ADMIN else {"tenant_id": current_user.tenant_id}
    # Include assets even without purchase_price so report is never empty
    assets = await db.assets.find({**tenant_q, "is_deleted": {"$ne": True}}, {"_id": 0}).to_list(1000)
    now = datetime.now(timezone.utc)
    result = []
    for a in assets:
        # Fall back to created_at if purchase_date not set
        purchase_date_str = a.get("purchase_date") or a.get("created_at")
        if not purchase_date_str:
            purchase_date_str = now.isoformat()
        try:
            purchase_date = datetime.fromisoformat(purchase_date_str.replace("Z", "+00:00"))
        except Exception:
            purchase_date = now
        years = max((now - purchase_date).days / 365.25, 0)
        price = a.get("purchase_price") or 0
        rate = (a.get("depreciation_rate") or 20) / 100
        salvage = a.get("salvage_value") or 0
        method = a.get("depreciation_method") or "straight_line"
        if method == "straight_line":
            annual_dep = (price - salvage) * rate
            current_val = max(salvage, price - annual_dep * years)
        elif method == "declining_balance":
            current_val = max(salvage, price * ((1 - rate) ** years))
        else:
            current_val = price
        result.append({
            "asset_tag": a.get("asset_tag"), "serial_number": a.get("serial_number"),
            "purchase_price": price, "current_value": round(current_val, 2),
            "depreciation": round(price - current_val, 2), "years_old": round(years, 1)
        })
    return result

# ═══════════════════════════════════════════════════════════════════════════
# ACTIVITY FEED
# ═══════════════════════════════════════════════════════════════════════════

@api_router.get("/activity-feed")
async def get_activity_feed(
    page: int = Query(1, ge=1),
    limit: int = Query(30, ge=1, le=100),
    action_type: Optional[str] = Query(None, description="Filter by action type"),
    resource_type: Optional[str] = Query(None, description="Filter by resource type"),
    date_from: Optional[str] = Query(None, description="Filter from date (ISO)"),
    date_to: Optional[str] = Query(None, description="Filter to date (ISO)"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    current_user: User = Depends(get_current_user_hybrid)
):
    skip = (page - 1) * limit
    query: dict = {}
    if action_type:
        query["action"] = action_type
    if resource_type:
        query["resource"] = resource_type
    if user_id:
        query["user_id"] = user_id
    if date_from or date_to:
        ts_filter: dict = {}
        if date_from:
            ts_filter["$gte"] = date_from
        if date_to:
            ts_filter["$lte"] = date_to
        query["timestamp"] = ts_filter
    activities = await db.audit_log.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return activities

# ═══════════════════════════════════════════════════════════════════════════
# RECURRING MAINTENANCE — when completing, auto-schedule next if interval set
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/maintenance/{maintenance_id}/complete")
async def complete_maintenance(
    maintenance_id: str,
    cost: float = 0.0,
    notes: str = "",
    recurrence_days: int = 0,
    current_user: User = Depends(get_current_user_hybrid)
):
    """Mark maintenance as completed and optionally schedule the next one."""
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    maint = await db.maintenance_schedules.find_one({"id": maintenance_id})
    if not maint:
        raise HTTPException(status_code=404, detail="Maintenance record not found")
    completed_date = datetime.now(timezone.utc).isoformat()
    await db.maintenance_schedules.update_one(
        {"id": maintenance_id},
        {"$set": {"status": "completed", "completed_date": completed_date, "cost": cost, "notes": notes}}
    )
    next_id = None
    if recurrence_days > 0:
        next_date = (datetime.now(timezone.utc) + timedelta(days=recurrence_days)).isoformat()
        next_maint = MaintenanceSchedule(
            asset_id=maint["asset_id"],
            tenant_id=maint["tenant_id"],
            title=maint["title"],
            description=maint.get("description", ""),
            scheduled_date=next_date,
            assigned_to=maint.get("assigned_to"),
            maintenance_type=maint.get("maintenance_type", "preventive"),
        )
        await db.maintenance_schedules.insert_one(next_maint.model_dump())
        next_id = next_maint.id
    return {"message": "Maintenance completed", "next_maintenance_id": next_id}

# ═══════════════════════════════════════════════════════════════════════════
# DATABASE INDEXES FOR NEW COLLECTIONS
# ═══════════════════════════════════════════════════════════════════════════
# (Indexes added in startup event below)

# Helper to check tier limits
async def check_tier_limit(tenant_id: str, resource: str):
    """Check if tenant has reached their tier limit for a resource. Raises HTTPException if exceeded."""
    if not tenant_id:
        return  # Super admin or no tenant
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        return
    
    tier_id = tenant.get("subscription_tier_id", "tier-free")
    tier = await db.subscription_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        return
    
    limits = tier.get("limits", {})
    
    if resource == "users":
        limit = limits.get("max_users", 3)
        if limit == -1:
            return
        count = await db.users.count_documents({"tenant_id": tenant_id})
        if count >= limit:
            raise HTTPException(status_code=403, detail=f"User limit reached ({limit}). Upgrade your subscription plan.")
    
    elif resource == "assets":
        limit = limits.get("max_assets", 10)
        if limit == -1:
            return
        count = await db.assets.count_documents({"tenant_id": tenant_id})
        if count >= limit:
            raise HTTPException(status_code=403, detail=f"Asset limit reached ({limit}). Upgrade your subscription plan.")
    
    elif resource == "orders":
        limit = limits.get("max_orders_per_month", 5)
        if limit == -1:
            return
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        count = await db.orders.count_documents({"tenant_id": tenant_id, "created_at": {"$gte": month_start.isoformat()}})
        if count >= limit:
            raise HTTPException(status_code=403, detail=f"Monthly order limit reached ({limit}). Upgrade your subscription plan.")
    
    elif resource == "tickets":
        limit = limits.get("max_tickets_per_month", 10)
        if limit == -1:
            return
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        count = await db.tickets.count_documents({"tenant_id": tenant_id, "created_at": {"$gte": month_start.isoformat()}})
        if count >= limit:
            raise HTTPException(status_code=403, detail=f"Monthly ticket limit reached ({limit}). Upgrade your subscription plan.")

# ═══════════════════════════════════════════════════════════════════════════
# FEATURE 13: DUPLICATE SERIAL NUMBER CHECK
# ═══════════════════════════════════════════════════════════════════════════

@api_router.get("/assets/check-serial")
async def check_serial_number(
    serial_number: str = Query(...),
    tenant_id: str = Query(...),
    current_user: User = Depends(get_current_user_hybrid)
):
    query = {"serial_number": {"$regex": f"^{re.escape(serial_number)}$", "$options": "i"}, "is_deleted": {"$ne": True}}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = tenant_id
    existing = await db.assets.find_one(query, {"_id": 0, "asset_tag": 1, "id": 1})
    if existing:
        return {"exists": True, "asset_tag": existing.get("asset_tag", ""), "asset_id": existing.get("id", "")}
    return {"exists": False}


# ═══════════════════════════════════════════════════════════════════════════
# FEATURE 3: ASSET DISPOSAL / WRITE-OFF
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/assets/{asset_id}/dispose")
async def dispose_asset(asset_id: str, data: AssetDispose, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if asset.get("status") == "disposed":
        raise HTTPException(status_code=400, detail="Asset is already disposed")
    update = {
        "status": "disposed",
        "disposal_date": data.disposal_date,
        "disposal_method": data.disposal_method,
        "disposal_notes": data.disposal_notes,
        "sale_proceeds": data.sale_proceeds,
    }
    await db.assets.update_one({"id": asset_id}, {"$set": update})
    history = AssetHistory(
        asset_id=asset_id, action="disposed",
        performed_by=current_user.id,
        notes=f"Disposed via {data.disposal_method}. Sale proceeds: {data.sale_proceeds}"
    )
    await db.asset_history.insert_one(history.model_dump())
    updated = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    return Asset(**updated)


# ═══════════════════════════════════════════════════════════════════════════
# FEATURE 2: ASSET DOCUMENT ATTACHMENTS
# ═══════════════════════════════════════════════════════════════════════════

ALLOWED_DOC_TYPES = {"application/pdf", "application/msword",
                     "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                     "image/jpeg", "image/png", "application/vnd.ms-excel",
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

@api_router.post("/assets/{asset_id}/documents")
async def upload_asset_document(
    asset_id: str,
    document_type: str = "other",
    notes: str = "",
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_hybrid)
):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if file.content_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported file type. Allowed: PDF, Word, Excel, JPEG, PNG")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File must be under 10MB")
    import base64
    b64 = base64.b64encode(content).decode()
    doc = AssetDocument(
        asset_id=asset_id,
        tenant_id=asset["tenant_id"],
        document_type=document_type,
        filename=file.filename or "document",
        file_data=f"data:{file.content_type};base64,{b64}",
        file_content_type=file.content_type,
        file_size=len(content),
        uploaded_by=current_user.id,
        notes=notes
    )
    await db.asset_documents.insert_one(doc.model_dump())
    return {"message": "Document uploaded", "id": doc.id, "filename": doc.filename, "document_type": doc.document_type}


@api_router.get("/assets/{asset_id}/documents")
async def get_asset_documents(asset_id: str, current_user: User = Depends(get_current_user_hybrid)):
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    docs = await db.asset_documents.find({"asset_id": asset_id}, {"_id": 0, "file_data": 0}).to_list(100)
    return docs


@api_router.get("/assets/{asset_id}/documents/{doc_id}/download")
async def download_asset_document(asset_id: str, doc_id: str, current_user: User = Depends(get_current_user_hybrid)):
    doc = await db.asset_documents.find_one({"id": doc_id, "asset_id": asset_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    import base64
    data_url = doc["file_data"]
    _, b64_data = data_url.split(",", 1)
    file_bytes = base64.b64decode(b64_data)
    from fastapi.responses import Response as FastAPIResponse
    return FastAPIResponse(
        content=file_bytes,
        media_type=doc["file_content_type"],
        headers={"Content-Disposition": f'attachment; filename="{doc["filename"]}"'}
    )


@api_router.delete("/assets/{asset_id}/documents/{doc_id}")
async def delete_asset_document(asset_id: str, doc_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.asset_documents.delete_one({"id": doc_id, "asset_id": asset_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document deleted"}


# ═══════════════════════════════════════════════════════════════════════════
# FEATURE 8: AMC CONTRACT TRACKING
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/assets/{asset_id}/amc")
async def create_amc_contract(asset_id: str, data: AMCContractCreate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    contract = AMCContract(asset_id=asset_id, tenant_id=asset["tenant_id"], **data.model_dump())
    await db.amc_contracts.insert_one(contract.model_dump())
    return contract


@api_router.get("/assets/{asset_id}/amc")
async def get_amc_contracts(asset_id: str, current_user: User = Depends(get_current_user_hybrid)):
    contracts = await db.amc_contracts.find({"asset_id": asset_id}, {"_id": 0}).to_list(50)
    return contracts


@api_router.patch("/amc/{contract_id}")
async def update_amc_contract(contract_id: str, data: AMCContractUpdate, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    update_data = data.model_dump(exclude_unset=True)
    result = await db.amc_contracts.update_one({"id": contract_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="AMC contract not found")
    return await db.amc_contracts.find_one({"id": contract_id}, {"_id": 0})


@api_router.delete("/amc/{contract_id}")
async def delete_amc_contract(contract_id: str, current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    result = await db.amc_contracts.delete_one({"id": contract_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="AMC contract not found")
    return {"message": "AMC contract deleted"}


# ═══════════════════════════════════════════════════════════════════════════
# FEATURE 6: EXCEL (.xlsx) EXPORT
# ═══════════════════════════════════════════════════════════════════════════

@api_router.get("/assets/export/xlsx")
async def export_assets_xlsx(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    query: dict = {"is_deleted": {"$ne": True}}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id
    assets = await db.assets.find(query, {"_id": 0}).to_list(5000)
    products_raw = await db.products.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
    products_map = {p["id"]: p["name"] for p in products_raw}

    import openpyxl
    import io
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = [
        "Asset Tag", "Product", "Serial Number", "Status", "Location",
        "Assigned To", "Purchase Date", "Purchase Price", "Current Value",
        "Depreciation Method", "Depreciation Rate (%)", "Warranty End Date",
        "Is Leased", "Lessor Name", "Lease End Date", "Monthly Payment",
        "Disposal Date", "Disposal Method", "Sale Proceeds",
        "Is Demo", "Created At"
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for a in assets:
        ws.append([
            a.get("asset_tag", ""),
            products_map.get(a.get("product_id", ""), "Unknown"),
            a.get("serial_number", ""),
            a.get("status", ""),
            a.get("location", ""),
            a.get("assigned_to", ""),
            a.get("purchase_date", "")[:10] if a.get("purchase_date") else "",
            a.get("purchase_price", 0),
            a.get("current_value", 0),
            a.get("depreciation_method", ""),
            a.get("depreciation_rate", 0),
            a.get("warranty_end_date", "")[:10] if a.get("warranty_end_date") else "",
            "Yes" if a.get("is_leased") else "No",
            a.get("lessor_name", ""),
            a.get("lease_end_date", "")[:10] if a.get("lease_end_date") else "",
            a.get("monthly_lease_payment", 0),
            a.get("disposal_date", "")[:10] if a.get("disposal_date") else "",
            a.get("disposal_method", ""),
            a.get("sale_proceeds", 0),
            "Yes" if a.get("is_demo") else "No",
            a.get("created_at", "")[:10] if a.get("created_at") else "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"assets_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ═══════════════════════════════════════════════════════════════════════════
# DEMO DATA SEEDING
# ═══════════════════════════════════════════════════════════════════════════

async def seed_tenant_demo_data(tenant_id: str, admin_user_id: str) -> int:
    """Seed comprehensive demo data for a new tenant. Tracks all seeded IDs in demo_data_registry."""
    now = datetime.now(timezone.utc)
    registry_entries = []

    def reg(collection: str, item_id: str):
        registry_entries.append({
            "tenant_id": tenant_id, "collection": collection,
            "item_id": item_id, "created_at": now.isoformat()
        })

    # Department
    dept_id = str(uuid.uuid4())
    await db.departments.insert_one({
        "id": dept_id, "name": "IT Department", "description": "Sample IT department for demo",
        "tenant_id": tenant_id, "budget": 500000.0, "created_at": now.isoformat()
    })
    reg("departments", dept_id)

    # Location
    loc_id = str(uuid.uuid4())
    await db.locations.insert_one({
        "id": loc_id, "name": "Head Office", "building": "Tower A",
        "floor": "3rd Floor", "room": "IT Lab", "tenant_id": tenant_id,
        "created_at": now.isoformat()
    })
    reg("locations", loc_id)

    # Vendors
    v1_id, v2_id = str(uuid.uuid4()), str(uuid.uuid4())
    await db.vendors.insert_many([
        {"id": v1_id, "name": "TechSource India", "contact_name": "Raj Kumar",
         "email": "raj@techsource.in", "phone": "+91-98765-43210",
         "category": "Hardware", "tenant_id": tenant_id,
         "notes": "Primary hardware supplier", "website": "https://techsource.in",
         "created_at": now.isoformat()},
        {"id": v2_id, "name": "SoftSolutions Ltd", "contact_name": "Priya Sharma",
         "email": "priya@softsolutions.com", "phone": "+91-98123-45678",
         "category": "Software", "tenant_id": tenant_id,
         "notes": "Software licensing partner", "website": "https://softsolutions.com",
         "created_at": now.isoformat()},
    ])
    reg("vendors", v1_id)
    reg("vendors", v2_id)

    # Products
    p_laptop_id = str(uuid.uuid4())
    p_monitor_id = str(uuid.uuid4())
    p_mouse_id = str(uuid.uuid4())
    await db.products.insert_many([
        {"id": p_laptop_id, "name": "Dell Latitude 5540", "category": "Laptop",
         "description": "Business laptop - 13th Gen Intel Core i5, 16GB RAM, 512GB SSD",
         "specs": {"processor": "Intel Core i5-1335U", "ram": "16GB", "storage": "512GB SSD", "display": "15.6 FHD"},
         "image_url": "", "stock": 5, "price": 85000.0, "tenant_id": tenant_id, "created_at": now.isoformat()},
        {"id": p_monitor_id, "name": "Samsung 24 Monitor", "category": "Monitor",
         "description": "24-inch Full HD IPS monitor with HDMI and VGA ports",
         "specs": {"size": "24 inch", "resolution": "1920x1080", "panel": "IPS"},
         "image_url": "", "stock": 8, "price": 18000.0, "tenant_id": tenant_id, "created_at": now.isoformat()},
        {"id": p_mouse_id, "name": "Logitech MX Master 3", "category": "Peripheral",
         "description": "Advanced wireless mouse with ergonomic design",
         "specs": {"connectivity": "Bluetooth + USB", "battery": "Rechargeable"},
         "image_url": "", "stock": 15, "price": 4500.0, "tenant_id": tenant_id, "created_at": now.isoformat()},
    ])
    reg("products", p_laptop_id)
    reg("products", p_monitor_id)
    reg("products", p_mouse_id)

    # Date helpers
    d180 = (now - timedelta(days=180)).isoformat()
    d365 = (now - timedelta(days=365)).isoformat()
    d730 = (now - timedelta(days=730)).isoformat()
    d90  = (now - timedelta(days=90)).isoformat()
    d30  = (now - timedelta(days=30)).isoformat()
    w365 = (now + timedelta(days=365)).isoformat()
    w540 = (now + timedelta(days=540)).isoformat()
    w1065 = (now + timedelta(days=1065)).isoformat()
    w275 = (now + timedelta(days=275)).isoformat()
    w180 = (now + timedelta(days=180)).isoformat()
    expired_w = (now - timedelta(days=365)).isoformat()

    def _base():
        return {
            "checked_out_to": None, "checkout_date": None,
            "expected_return_date": None, "actual_return_date": None,
            "is_leased": False, "lessor_name": "",
            "lease_start_date": None, "lease_end_date": None, "monthly_lease_payment": 0.0,
            "disposal_date": None, "disposal_method": "", "disposal_notes": "", "sale_proceeds": 0.0,
            "is_deleted": False, "department_id": dept_id,
        }

    assets = [
        {**_base(), "id": str(uuid.uuid4()), "asset_tag": "ASSET-001",
         "product_id": p_laptop_id, "serial_number": "DELL-SN-001234",
         "tenant_id": tenant_id, "status": "assigned",
         "assigned_to": admin_user_id, "assigned_date": d180,
         "location": "Head Office", "purchase_date": d365,
         "warranty_start_date": d365, "warranty_end_date": w365,
         "warranty_provider": "Dell India", "purchase_price": 85000.0,
         "depreciation_method": "straight_line", "depreciation_rate": 20.0,
         "salvage_value": 5000.0, "current_value": 72000.0,
         "is_demo": True, "expiry_date": None, "created_at": d365},
        {**_base(), "id": str(uuid.uuid4()), "asset_tag": "ASSET-002",
         "product_id": p_monitor_id, "serial_number": "SAM-MON-88421",
         "tenant_id": tenant_id, "status": "available",
         "assigned_to": None, "assigned_date": None,
         "location": "IT Lab", "purchase_date": d180,
         "warranty_start_date": d180, "warranty_end_date": w540,
         "warranty_provider": "Samsung India", "purchase_price": 18000.0,
         "depreciation_method": "straight_line", "depreciation_rate": 20.0,
         "salvage_value": 1000.0, "current_value": 15500.0,
         "is_demo": True, "expiry_date": None, "created_at": d180},
        {**_base(), "id": str(uuid.uuid4()), "asset_tag": "ASSET-003",
         "product_id": p_laptop_id, "serial_number": "DELL-SN-005678",
         "tenant_id": tenant_id, "status": "under_maintenance",
         "assigned_to": None, "assigned_date": None,
         "location": "Service Center", "purchase_date": d730,
         "warranty_start_date": d730, "warranty_end_date": expired_w,
         "warranty_provider": "Dell India", "purchase_price": 78000.0,
         "depreciation_method": "straight_line", "depreciation_rate": 20.0,
         "salvage_value": 5000.0, "current_value": 46800.0,
         "is_demo": True, "expiry_date": None, "created_at": d730},
        {**_base(), "id": str(uuid.uuid4()), "asset_tag": "ASSET-004",
         "product_id": p_mouse_id, "serial_number": "LOG-MX3-99012",
         "tenant_id": tenant_id, "status": "available",
         "assigned_to": None, "assigned_date": None,
         "location": "Storage Room", "purchase_date": d90,
         "warranty_start_date": d90, "warranty_end_date": w275,
         "warranty_provider": "Logitech", "purchase_price": 4500.0,
         "depreciation_method": "straight_line", "depreciation_rate": 25.0,
         "salvage_value": 200.0, "current_value": 4192.0,
         "is_demo": True, "expiry_date": None, "created_at": d90},
        {**_base(), "id": str(uuid.uuid4()), "asset_tag": "ASSET-005",
         "product_id": p_laptop_id, "serial_number": "DELL-SN-007890",
         "tenant_id": tenant_id, "status": "available",
         "assigned_to": None, "assigned_date": None,
         "location": "Head Office", "purchase_date": d30,
         "warranty_start_date": d30, "warranty_end_date": w1065,
         "warranty_provider": "Dell India", "purchase_price": 92000.0,
         "depreciation_method": "straight_line", "depreciation_rate": 20.0,
         "salvage_value": 5000.0, "current_value": 91500.0,
         "is_demo": True, "expiry_date": None, "created_at": d30},
    ]
    await db.assets.insert_many(assets)
    for a in assets:
        reg("assets", a["id"])

    # Tickets
    t_count = await db.tickets.count_documents({"tenant_id": tenant_id})
    tickets = [
        {"id": str(uuid.uuid4()),
         "ticket_number": f"TKT-{str(t_count + 1001).zfill(4)}",
         "tenant_id": tenant_id, "created_by": admin_user_id, "assigned_to": None,
         "title": "Laptop screen flickering on ASSET-001",
         "description": "The screen on the assigned Dell Latitude is flickering intermittently during video calls.",
         "priority": "high", "status": "open", "category": "hardware",
         "created_at": (now - timedelta(days=2)).isoformat(),
         "updated_at": (now - timedelta(days=2)).isoformat()},
        {"id": str(uuid.uuid4()),
         "ticket_number": f"TKT-{str(t_count + 1002).zfill(4)}",
         "tenant_id": tenant_id, "created_by": admin_user_id, "assigned_to": None,
         "title": "VPN not connecting from home office",
         "description": "Unable to connect to company VPN from home. Error: Authentication failed.",
         "priority": "medium", "status": "in_progress", "category": "network",
         "created_at": (now - timedelta(days=5)).isoformat(),
         "updated_at": (now - timedelta(days=1)).isoformat()},
        {"id": str(uuid.uuid4()),
         "ticket_number": f"TKT-{str(t_count + 1003).zfill(4)}",
         "tenant_id": tenant_id, "created_by": admin_user_id, "assigned_to": None,
         "title": "Printer offline on 2nd floor",
         "description": "HP LaserJet printer showing offline. Multiple users unable to print.",
         "priority": "low", "status": "resolved", "category": "hardware",
         "created_at": (now - timedelta(days=10)).isoformat(),
         "updated_at": (now - timedelta(days=7)).isoformat()},
    ]
    await db.tickets.insert_many(tickets)
    for t in tickets:
        reg("tickets", t["id"])

    # Licenses
    lic1_id = str(uuid.uuid4())
    lic2_id = str(uuid.uuid4())
    await db.licenses.insert_many([
        {"id": lic1_id, "name": "Microsoft 365 Business", "vendor": "Microsoft",
         "license_key": "XXXXX-XXXXX-XXXXX-XXXXX-XXXXX",
         "seats_total": 25, "seats_used": 18,
         "purchase_date": d365, "expiry_date": w365,
         "cost": 85000.0, "license_type": "subscription",
         "tenant_id": tenant_id, "notes": "Annual subscription for office productivity",
         "created_at": d365},
        {"id": lic2_id, "name": "Adobe Creative Cloud", "vendor": "Adobe",
         "license_key": "ADOBE-CC-DEMO-2024",
         "seats_total": 5, "seats_used": 3,
         "purchase_date": d180, "expiry_date": w180,
         "cost": 32000.0, "license_type": "subscription",
         "tenant_id": tenant_id, "notes": "Creative Cloud for design team",
         "created_at": d180},
    ])
    reg("licenses", lic1_id)
    reg("licenses", lic2_id)

    # Order
    ord_id = str(uuid.uuid4())
    await db.orders.insert_one({
        "id": ord_id, "tenant_id": tenant_id, "user_id": admin_user_id,
        "product_id": p_laptop_id, "quantity": 2,
        "delivery_notes": "Deliver to IT Lab, Tower A",
        "status": "pending", "approval_stage": None,
        "checker_approved_by": None, "checker_approved_date": None,
        "approver_approved_by": None, "approver_approved_date": None,
        "approved_by": None, "approval_date": None,
        "rejection_reason": "",
        "created_at": (now - timedelta(days=3)).isoformat()
    })
    reg("orders", ord_id)

    if registry_entries:
        await db.demo_data_registry.insert_many(registry_entries)

    return len(registry_entries)


@api_router.get("/demo-data/status")
async def get_demo_data_status(current_user: User = Depends(get_current_user_hybrid)):
    tenant_id = current_user.tenant_id
    if not tenant_id:
        return {"has_demo_data": False, "count": 0}
    count = await db.demo_data_registry.count_documents({"tenant_id": tenant_id})
    return {"has_demo_data": count > 0, "count": count}


@api_router.post("/demo-data/seed")
async def reseed_demo_data(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="No tenant associated with your account")
    existing = await db.demo_data_registry.count_documents({"tenant_id": current_user.tenant_id})
    if existing > 0:
        raise HTTPException(status_code=400, detail="Demo data already exists. Delete it first before re-seeding.")
    count = await seed_tenant_demo_data(current_user.tenant_id, current_user.id)
    return {"message": "Demo data seeded successfully", "items_created": count}


@api_router.delete("/demo-data")
async def delete_demo_data(current_user: User = Depends(get_current_user_hybrid)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    tenant_id = current_user.tenant_id
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No tenant associated with your account")
    registry = await db.demo_data_registry.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(2000)
    by_collection: dict = {}
    for entry in registry:
        col = entry["collection"]
        by_collection.setdefault(col, []).append(entry["item_id"])
    deleted_count = 0
    for collection_name, ids in by_collection.items():
        result = await db[collection_name].delete_many({"id": {"$in": ids}})
        deleted_count += result.deleted_count
    await db.demo_data_registry.delete_many({"tenant_id": tenant_id})
    await write_audit_log(current_user.id, "delete", "demo_data", tenant_id, f"Cleared {deleted_count} demo items")
    return {"message": "Demo data cleared successfully", "items_deleted": deleted_count}


# ═══════════════════════════════════════════════════════════════════════════
# BULK ASSET IMPORT (CSV / Excel)
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/assets/import")
async def import_assets(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN, UserRole.ASSET_MANAGER]:
        raise HTTPException(status_code=403, detail="Access denied")
    tenant_id = current_user.tenant_id or "default"
    content = await file.read()
    rows = []
    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".csv"):
            import io, csv as csv_module
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv_module.DictReader(io.StringIO(text))
            rows = list(reader)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            import io, openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h is not None else "" for h in row]
                    continue
                rows.append(dict(zip(headers, [str(c).strip() if c is not None else "" for c in row])))
        else:
            raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    # Pre-load products for name lookup — super_admin sees all tenants' products
    if current_user.role == UserRole.SUPER_ADMIN:
        products_list = await db.products.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1, "tenant_id": 1}).to_list(5000)
    else:
        products_list = await db.products.find({"tenant_id": tenant_id, "is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1, "tenant_id": 1}).to_list(2000)
    product_map = {p["name"].strip().lower(): p["id"] for p in products_list}

    created, skipped = 0, []
    for i, row in enumerate(rows, start=2):
        asset_tag = str(row.get("asset_tag") or row.get("Asset Tag") or "").strip()
        serial_number = str(row.get("serial_number") or row.get("Serial Number") or "").strip()
        if not asset_tag or not serial_number:
            skipped.append(f"Row {i}: missing asset_tag or serial_number")
            continue
        # Resolve product_id — accept many common column name variations
        product_id = str(
            row.get("product_id") or row.get("Product ID") or row.get("ProductID") or ""
        ).strip()
        if not product_id:
            pname = str(
                row.get("product_name") or row.get("Product Name") or
                row.get("product") or row.get("Product") or
                row.get("ProductName") or ""
            ).strip().lower()
            product_id = product_map.get(pname, "")
        if not product_id:
            skipped.append(f"Row {i}: product '{pname}' not found — check product_name matches your product catalog")
            continue
        # Check duplicate serial
        existing = await db.assets.find_one({"serial_number": serial_number, "tenant_id": tenant_id})
        if existing:
            skipped.append(f"Row {i}: serial {serial_number} already exists")
            continue
        def _float(v):
            try: return float(v)
            except: return 0.0
        asset = Asset(
            asset_tag=asset_tag,
            product_id=product_id,
            serial_number=serial_number,
            tenant_id=tenant_id,
            location=str(row.get("location") or row.get("Location") or "").strip(),
            purchase_date=str(row.get("purchase_date") or row.get("Purchase Date") or "").strip() or None,
            warranty_start_date=str(row.get("warranty_start_date") or "").strip() or None,
            warranty_end_date=str(row.get("warranty_end_date") or row.get("Warranty End Date") or "").strip() or None,
            warranty_provider=str(row.get("warranty_provider") or row.get("Warranty Provider") or "").strip(),
            purchase_price=_float(row.get("purchase_price") or row.get("Purchase Price") or 0),
            department_id=str(row.get("department_id") or "").strip() or None,
        )
        await db.assets.insert_one(asset.model_dump())
        created += 1

    return {"created": created, "skipped": len(skipped), "errors": skipped}


@api_router.get("/assets/import/template")
async def download_import_template(current_user: User = Depends(get_current_user)):
    """Return a CSV template for bulk asset import."""
    import io, csv as csv_module
    output = io.StringIO()
    writer = csv_module.writer(output)
    writer.writerow(["asset_tag", "serial_number", "product_name", "location", "purchase_date",
                     "warranty_end_date", "warranty_provider", "purchase_price"])
    writer.writerow(["AST-001", "SN-123456", "Dell Laptop", "Head Office", "2024-01-15",
                     "2027-01-15", "Dell Inc", "75000"])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=asset_import_template.csv"}
    )


# ═══════════════════════════════════════════════════════════════════════════
# USER INVITATION SYSTEM
# ═══════════════════════════════════════════════════════════════════════════

@api_router.post("/auth/invite")
async def create_invite(invite_data: InviteCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    # For non-super-admin, force their own tenant
    tenant_id = invite_data.tenant_id if current_user.role == UserRole.SUPER_ADMIN else current_user.tenant_id
    # Check if user with this email already exists
    existing_user = await db.users.find_one({"email": invite_data.email.lower()})
    if existing_user:
        raise HTTPException(status_code=400, detail="A user with this email already exists")
    # Invalidate any existing pending invites for this email
    await db.invite_tokens.delete_many({"email": invite_data.email.lower(), "accepted": False})
    invite = InviteToken(
        email=invite_data.email.lower(),
        role=invite_data.role,
        tenant_id=tenant_id,
        invited_by=current_user.id
    )
    await db.invite_tokens.insert_one(invite.model_dump())
    # Send invite email
    frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
    invite_link = f"{frontend_url}/invite/{invite.token}"
    smtp_host = os.environ.get("SMTP_HOST", "")
    if smtp_host:
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = "You're invited to join the Asset Management System"
            msg["From"] = os.environ.get("SMTP_FROM", "noreply@assetmanagement.com")
            msg["To"] = invite.email
            html_body = f"""
            <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
              <h2 style="color:#4F46E5">You're Invited!</h2>
              <p>{current_user.name} has invited you to join as <strong>{invite_data.role.replace('_',' ').title()}</strong>.</p>
              <p style="margin:24px 0">
                <a href="{invite_link}" style="background:#4F46E5;color:#fff;padding:12px 24px;border-radius:6px;text-decoration:none;font-weight:bold">Accept Invitation</a>
              </p>
              <p style="color:#888;font-size:12px">This link expires in 7 days. If you did not expect this invite, please ignore this email.</p>
            </div>"""
            msg.attach(MIMEText(html_body, "html"))
            smtp_port = int(os.environ.get("SMTP_PORT", 587))
            smtp_user = os.environ.get("SMTP_USER", "")
            smtp_pass = os.environ.get("SMTP_PASS", "")
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            if smtp_user:
                server.login(smtp_user, smtp_pass)
            server.sendmail(msg["From"], [invite.email], msg.as_string())
            server.quit()
        except Exception as e:
            logging.warning(f"Failed to send invite email: {e}")
    return {"message": "Invitation created", "token": invite.token, "invite_link": invite_link}


@api_router.get("/auth/invite/{token}")
async def get_invite(token: str):
    invite = await db.invite_tokens.find_one({"token": token, "accepted": False}, {"_id": 0})
    if not invite:
        raise HTTPException(status_code=404, detail="Invite not found or already used")
    if datetime.fromisoformat(invite["expires_at"]) < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="Invite link has expired")
    return {"email": invite["email"], "role": invite["role"], "tenant_id": invite.get("tenant_id")}


@api_router.post("/auth/invite/{token}/accept")
async def accept_invite(token: str, accept_data: InviteAccept):
    invite = await db.invite_tokens.find_one({"token": token, "accepted": False}, {"_id": 0})
    if not invite:
        raise HTTPException(status_code=404, detail="Invite not found or already used")
    if datetime.fromisoformat(invite["expires_at"]) < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="Invite link has expired")
    validate_password_strength(accept_data.password)
    # Check email not already taken (race condition guard)
    if await db.users.find_one({"email": invite["email"]}):
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed = pwd_context.hash(accept_data.password)
    user = User(
        email=invite["email"],
        name=accept_data.name,
        hashed_password=hashed,
        role=invite["role"],
        tenant_id=invite.get("tenant_id"),
        status="active"
    )
    await db.users.insert_one(user.model_dump())
    await db.invite_tokens.update_one({"token": token}, {"$set": {"accepted": True}})
    return {"message": "Account created successfully. You can now log in."}


@api_router.get("/auth/invites")
async def list_invites(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    query: dict = {"accepted": False}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id
    invites = await db.invite_tokens.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return invites


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_db_client():
    """Create database indexes for fast queries."""
    await db.users.create_index("email", unique=True, background=True)
    await db.users.create_index("tenant_id", background=True)
    await db.assets.create_index("tenant_id", background=True)
    await db.assets.create_index("status", background=True)
    await db.assets.create_index("assigned_to", background=True)
    await db.products.create_index("tenant_id", background=True)
    await db.orders.create_index("tenant_id", background=True)
    await db.orders.create_index("status", background=True)
    await db.orders.create_index("user_id", background=True)
    await db.tickets.create_index("tenant_id", background=True)
    await db.tickets.create_index("status", background=True)
    await db.tickets.create_index("created_by", background=True)
    await db.audit_log.create_index("timestamp", background=True)
    await db.departments.create_index("tenant_id", background=True)
    await db.asset_history.create_index("asset_id", background=True)
    await db.locations.create_index("tenant_id", background=True)
    await db.vendors.create_index("tenant_id", background=True)
    await db.licenses.create_index("tenant_id", background=True)
    await db.reservations.create_index("tenant_id", background=True)
    await db.reservations.create_index("reserved_by", background=True)
    await db.ticket_comments.create_index("ticket_id", background=True)
    await db.asset_transfers.create_index("tenant_id", background=True)
    await db.custom_fields.create_index("tenant_id", background=True)
    await db.api_keys.create_index("user_id", background=True)
    await db.api_keys.create_index("key", unique=True, background=True)
    await db.invoices.create_index("tenant_id", background=True)
    await db.invoices.create_index("created_at", background=True)
    await db.asset_documents.create_index("asset_id", background=True)
    await db.amc_contracts.create_index("asset_id", background=True)
    await db.amc_contracts.create_index("tenant_id", background=True)
    await db.demo_data_registry.create_index("tenant_id", background=True)
    await db.invite_tokens.create_index("token", background=True)
    await db.invite_tokens.create_index("email", background=True)
    await seed_default_tiers()
    asyncio.create_task(run_scheduled_alerts())
    logger.info("Database indexes created, default tiers seeded, and alert scheduler started.")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=True)